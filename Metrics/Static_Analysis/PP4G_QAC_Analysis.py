###################################
# For more information, see " EXXXXXX_SBE_PP4G_HowTo_PassQACAnalysis.docx " located on S:\Methods\HowTo
## QA·C references the following environment variables:
## QACBIN : The location of QA·C’s system files such as the message file, qac.msg, and the configuration file, qac.cfg. This should be set to the bin directory.
## QACHELPFILES : The location of the HTML help pages, explaining the rationale behind each generated message. This should be set to the help directory, and message HTML help files located in \messages sub-directory.
## QACOUTPATH : The location in which output files are generated. This is usually part of project configuration, and, in the GUI, is set in the folder parameters. 
## 
##Note: If you intend to run QA·C concurrently (e.g. GUI and qac.exe), it is unsafe to rely on QACOUTPATH. It is recommended to specify the output path with an -op entry, instead.
###################################

import os
import argparse
import subprocess
import re
import openpyxl
import datetime
import shutil
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors, GradientFill
from openpyxl import load_workbook
from pathlib import Path
import difflib
import sys
import io
import glob
import fnmatch
import ntpath
import csv
import warnings
import win32api
from win32com import client

def findLineMarker(inputLine):  
    matchRet = re.match("(\*\*\* )([0-9]+),([0-9]+)( \*\*\*\*)", inputLine)
    if matchRet is None:
        matchRet = re.match("(--- )([0-9]+),([0-9]+)( ----)", inputLine)
    return matchRet

def run_command_str(commandStr):
    retStr = ''
    p = subprocess.Popen(commandStr, stdout=subprocess.PIPE, stderr=subprocess.PIPE, )

    for line in io.TextIOWrapper(p.stdout, encoding="utf-8"):
        line = line.strip()
        if line:
            retStr += '{}\n'.format(line)

    for line in io.TextIOWrapper(p.stderr, encoding="utf-8"):
        line = line.strip()
        if line:
            retStr += '{}\n'.format(line)

    p.wait()

    return retStr, p.returncode
	
def createViaFile(configFileList):
    global workingFolder
    viaFileLocation = os.path.join(workingFolder, "settings.via")
    with open(viaFileLocation, "wt", errors='ignore') as viaFile:
        for file in configFiles:
            with open(file, "rt", errors='ignore') as fileContent:
                for line in fileContent:
                    viaFile.write(line)

def analyzeCFile(fileFolder, fileName):
    global QAC_folder, workingFolder
    command = [os.path.join(QAC_folder, "bin", "qac.exe")]
    viaFileLocation = os.path.join(workingFolder, "settings.via")
    command += ["-via", viaFileLocation]
    command += ["-op", workingFolder]
    command += [os.path.join(fileFolder, fileName)]

    cmd_output, cmd_error_code = run_command_str(" ".join(command))
    if cmd_error_code != 0:
        print('qac.exe returned code: ' + str(cmd_error_code))
        if cmd_output.strip():
            print(cmd_output)
        sys.exit()


    #Check existence of .err and .i and .met files
    l = [fileName + ".err", fileName + ".i", fileName + ".met"]
    for f in l:
        if not os.path.isfile(os.path.join(workingFolder, f)):
            return False
    return True

def clearTempFiles(fileName):
    l = [fileName + ".err", fileName + ".i", fileName + ".met", fileName + "_warning.txt"]
    for f in l:
        name = os.path.join(workingFolder, f)
        if os.path.isfile(name):
            os.remove(name)

def clearTempFiles(fileName):
    l = [fileName + ".err", fileName + ".i", fileName + ".met", fileName + "_warning.txt"]
    for f in l:
        name = os.path.join(workingFolder, f)
        if os.path.isfile(name):
            os.remove(name)


def generateHtmlReport(fileFolder, fileName):
    global QAC_folder, workingFolder
    command = [os.path.join(QAC_folder, "bin", "errdsp.exe"), "QAC"]
    viaFileLocation = os.path.join(workingFolder, "settings.via")
    command += ["-via", viaFileLocation]
    command += ["-op", workingFolder]
    command += ["-html"]
    command += ["-file", fileName + ".html"]
    command += [os.path.join(fileFolder, fileName)]

    cmd_output, cmd_error_code = run_command_str(" ".join(command))
    if cmd_error_code != 0:
        print('errdsp.exe returned code: ' + str(cmd_error_code))
        if cmd_output.strip():
            print(cmd_output)

def extractWarningsToTxt(fileFolder, fileName):
    global QAC_folder, workingFolder
    command = [os.path.join(QAC_folder, "bin", "errdsp.exe"), "QAC"]
    viaFileLocation = os.path.join(workingFolder, "settings.via")
    command += ["-via", viaFileLocation]
    command += ["-op", workingFolder]
    command += ["-file", fileName + "_warning.txt"]
    command += ["-m"]
    command += [os.path.join(fileFolder, fileName)]

    cmd_output, cmd_error_code = run_command_str(" ".join(command))
    if cmd_error_code != 0:
        print('errdsp.exe returned code: ' + str(cmd_error_code))
        if cmd_output.strip():
            print(cmd_output)

def extractWarningsToList(fileName):
    global workingFolder
    f = os.path.join(workingFolder, fileName + "_warning.txt")
    ret = []
    with open(f, "rt", errors='ignore') as file:
        regex = re.compile(r".*\\" + re.escape(fileName) + r" *\((\d+)\) *\+\+ ([ \w]+) \+\+: <=(\d+)=\((\d+)\) (.*)")
        otherRegex = re.compile(r"(.*)\((\d+)\) *\+\+ ([ \w]+) \+\+: <=(\d+)=\((\d+)\) (.*)")
        for line in file:
            m = regex.search(line)
            if m:
                d = dict()
                d['line_in_src'] = int(m.group(1))
                d['error_warning'] = m.group(2)
                d['severity'] = int(m.group(3))
                d['code'] = int(m.group(4))
                d['message'] = m.group(5)
                ret.append(d)
                #print(str(severity) + " (" + code +") at line " + str(line_in_src))
            else:
                n = otherRegex.search(line)
                if n:
                    if n.group(4) == "8" or n.group(4) == "9":
                        print("hard error detected. please check preprocessor")
                        print(n.group(0))


    return ret

def readSrcFile(fileFolder, fileName):
    filePath = os.path.join(fileFolder, fileName)
    res = []
    with open(filePath, "rt", errors='ignore') as file:
        for line in file:
            res.append(line)
    return res

def getPTCFileVersion(srcFile):
    # regex_rev  = ".*Current revision: *\$Rev"
    # regex_rev += "ision: *(.*).*\$.*"
    for line in srcFile:
        if '$Revision:' in line:
            version = line.split('Revision:')[1].split('$')[0].replace(' ', '')
            return version
        # m = re.match(regex_rev, line)
        # if m:
        #     return m.group(1).strip()
    return "Unknown"

def searchPreviousCommentOpening(srcFile, lineStart):
    for idx, line in enumerate(srcFile[lineStart::-1]):
        if line.strip().startswith("/*"):
            return lineStart - idx
    return None

def searchNextCommentOpening(srcFile, lineStart):
    for idx, line in enumerate(srcFile[lineStart:]):
        if line.strip().startswith("/*"):
            return idx + lineStart
    return None

def searchNextCommentClosing(srcFile, lineStart):
    for idx, line in enumerate(srcFile[lineStart:]):
        if line.strip().endswith("*/"):
            return idx + lineStart
    return None

def searchNextLineOfCodeNotEmpty(srcFile, lineStart, isInComment = True):
    if isInComment:
        lineStart = searchNextCommentClosing(srcFile, lineStart)
        if lineStart == None:
            raise Exception("Cannot find end of comment starting at line " + str(lineStart+1))
        lineStart += 1

    running = True
    while running:
        # lineStart is after last comment closing
        nextCommentOpeningLine = searchNextCommentOpening(srcFile, lineStart)
        for idx, linevalue in enumerate(srcFile[lineStart:nextCommentOpeningLine]):
            if linevalue.strip() != '':
                return lineStart + idx
        # If we're here, then we didn't find any source code in [lineStart ; nextCommentOpeningLine [
        if nextCommentOpeningLine is None:
            # No new comment opening
            return -1
        lineStart = searchNextCommentClosing(srcFile, nextCommentOpeningLine)
        if lineStart == None:
            raise Exception("Cannot find end of comment starting at line " + str(nextCommentOpeningLine+1))
        lineStart += 1


def searchForPRQA(srcFile, fileName, fileFolder):
    ret = numberOfPRQA = 0
    #filePath = os.path.join(fileFolder, fileName)
    for idx, line in enumerate(srcFile):
        if "PRQA" in line:
            numberOfPRQA += 1
            # Replace by QAC_WARNING
            #tempFile = open(filePath, 'r+')
            #tempFile.write(line.replace("PRQA","QAC_WARNING"))
            #tempFile.close()
    if numberOfPRQA != 0:
        print("/****************************************************************/")
        print("/*             " + str(fileName)  + "                         */")
        print("/* BE CAREFUL : " + str(numberOfPRQA) + " PRQA detected                */")
        print("/* Before QAC analysis, all PRQAs have to be removed.           */")
        print("/****************************************************************/")
    return ret


def searchForQACJustification(srcFile):
    ret = []
    for idx, line in enumerate(srcFile):
        if "QAC ID:" in line:
            m = re.match(" *QAC ID: *(\d+).*", line)
            if m:
                d = dict()
                code = int(m.group(1))
                d['code'] = code
                startCommentLine = searchPreviousCommentOpening(srcFile, idx)
                endCommentLine = searchNextCommentClosing(srcFile, startCommentLine)
                d['justification'] = srcFile[startCommentLine:endCommentLine+1]
                d['justifyLine'] = searchNextLineOfCodeNotEmpty(srcFile, idx) + 1
                ret.append(d)
    return ret

def addSrcFileInformationToWarningsList(srcFile, warningList):
    for elem in warningList:
        line = elem['line_in_src'] - 1
        elem['src_code'] = srcFile[line].strip()

def addQACJustificationToWarningsList(warningList, QACJustifications, margin = 2):
    #A warning contains ['line_in_src'], ['error_warning'], ['severity'], ['code'], ['message'], ['src_code']
    #A justification contains ['code'], ['justification'], ['justifyLine']
    #If a justification is found, warning will contains ['justification'] entry

    def passe(warningList, warningListRemainingIdx, QACJustifications, QACJustificationsRemainingIdx, margin = 0):
        for j_idx, justification in enumerate(QACJustifications):
            if j_idx in QACJustificationsRemainingIdx:
                #Try to find if a warning with the same code and same line match
                code = justification['code']
                line = justification['justifyLine']
                for w_idx, warning in enumerate(warningList):
                    #Justify line can be less than the warning line in case of multiple line statement
                    if warning['code'] == code and warning['line_in_src'] >= line and warning['line_in_src'] <= line + margin:
                        # Match
                        warning['justification'] = justification
                        if j_idx in QACJustificationsRemainingIdx:
                            QACJustificationsRemainingIdx.remove(j_idx)
                        if w_idx in warningListRemainingIdx:
                            warningListRemainingIdx.remove(w_idx)

    warningListRemainingIdx = [e for e in range(len(warningList))]
    QACJustificationsRemainingIdx = [e for e in range(len(QACJustifications))]
    passe(warningList, warningListRemainingIdx, QACJustifications, QACJustificationsRemainingIdx)
    passe(warningList, warningListRemainingIdx, QACJustifications, QACJustificationsRemainingIdx, margin = margin)


def generateWarningReport(warningList, fileName, fileVersion):
    global workingFolder
    global swProject
    section = 0

    f = os.path.join(workingFolder, fileName + "_WarningReport.xlsx")
    sheet_name = "Analysis of " + fileName

    old_fileName = fileName + "_WarningReport_Old.xlsx"
    old_path = os.path.join(workingFolder, old_fileName)

    # Make a copy of the old xlsx
    if os.path.isfile(f):
        shutil.copy(f, old_path)
        old_wb = load_workbook(os.path.join(workingFolder, old_fileName), data_only=True)
        old_sh = old_wb.worksheets[0]

    # Prepare the new xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[0:31]

    # CONTENT for header cells
    ws['A1'] = "Project " + swProject
    ws['A2'] = "Warning report of " + fileName + " " + fileVersion
    ws['A3'] = "Generated at " + str(datetime.datetime.now())
    ws['A4'] = "Line"
    ws['B4'] = "Type"
    ws['C4'] = "Severity"
    ws['D4'] = "Warning code"
    ws['E4'] = "Warning message"
    #ws['F3'] = "Justified"
    ws['F4'] = "Source"
    ws['G4'] = "Justification"
    ws['H4'] = "Action"
    ws['I4'] = "Status"

    # DIMENSION for columns
    ws.column_dimensions['A'].width  = 6
    ws.column_dimensions['C'].width  = 6
    ws.column_dimensions['D'].width  = 6
    ws.column_dimensions['E'].width  = 50
    ws.column_dimensions['F'].width  = 50
    ws.column_dimensions['G'].width  = 40
    ws.column_dimensions['H'].width  = 40
    ws.column_dimensions['I'].width  = 18

    # BORDERS for header cells
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    bottom = Border(bottom=border.bottom)
    right = Border(right=border.right)
    top = Border(top=border.top)
    square = right + bottom
    more = square + top
    ws['A1'].border = bottom
    ws['B1'].border = bottom
    ws['C1'].border = bottom
    ws['D1'].border = bottom
    ws['E1'].border = square
    ws['E2'].border = right
    ws['A3'].border = bottom
    ws['B3'].border = bottom
    ws['C3'].border = bottom
    ws['D3'].border = bottom
    ws['E3'].border = square
    ws['A4'].border = square
    ws['B4'].border = square
    ws['C4'].border = square
    ws['D4'].border = square
    ws['E4'].border = square
    ws['F4'].border = more
    ws['G4'].border = more
    ws['H4'].border = more
    ws['I4'].border = more

    # BACKGROUND COLORS for header cells
    color_generated_cells = "C0C0C0"
    color_manual_cells = "C4C4C4"
    fill_generated_cells = PatternFill("solid", fgColor=color_generated_cells)
    fill_manual_cells = PatternFill("solid", fgColor=color_manual_cells)
    for index_row in range(1, 5):
        for index_column in range(1, 6):
            ws.cell(row=index_row, column=index_column).fill = fill_generated_cells
    ws['F4'].fill = fill_generated_cells
    ws['G4'].fill = fill_manual_cells
    ws['H4'].fill = fill_manual_cells
    ws['I4'].fill = fill_manual_cells

    orig_file_str = ''
    backup_file_str = ''
    idx_of_printed_warning = 0
    for idx, w in enumerate(warningList):
        if warningList[idx]['severity'] > 4:
            row = 5 + idx_of_printed_warning
            idx_of_printed_warning += 1
            ws.cell(row = row, column = 1).value = w['line_in_src']
            ws.cell(row = row, column = 2).value = w['error_warning']
            ws.cell(row = row, column = 3).value = w['severity']
            ws.cell(row = row, column = 4).value = w['code']
            ws.cell(row = row, column = 5).value = w['message']
            ws.cell(row = row, column = 5).alignment = openpyxl.styles.Alignment(wrap_text=True)
            #ws.cell(row = row, column = 6).value = "TRUE" if 'justification' in w else "FALSE"
            ws.cell(row = row, column = 6).value = w['src_code']
            ws.cell(row = row, column = 6).alignment = openpyxl.styles.Alignment(wrap_text=True)
            #if 'justification' in w:
            #    st = ''
            #    for s in w['justification']['justification']:
            #        st += s
            #    ws.cell(row = row, column = 8).value = st.strip()
            #    ws.cell(row = row, column = 8).alignment = openpyxl.styles.Alignment(wrap_text=True)
            #else:
            #    ws.cell(row = row, column = 8).value = ""
            ws.cell(row = row, column = 1).border = square
            ws.cell(row = row, column = 2).border = square
            ws.cell(row = row, column = 3).border = square
            ws.cell(row = row, column = 4).border = square
            ws.cell(row = row, column = 5).border = square
            ws.cell(row = row, column = 6).border = square
            ws.cell(row = row, column = 7).border = square
            ws.cell(row = row, column = 8).border = square
            ws.cell(row = row, column = 9).border = square

            orig_file_str = os.path.normpath(os.path.join(workingFolder, "../../Implementation/src", fileName))
            if not os.path.isfile(orig_file_str):
                orig_file_str = os.path.normpath(os.path.join(workingFolder, "../../Implementation/cfg", fileName))
            backup_file_str = os.path.normpath(os.path.join(workingFolder, Path(fileName).stem + "_Old" + Path(fileName).suffix))

            diff_list = []

            if os.path.exists(backup_file_str):
                orig_file = open(orig_file_str, 'r', errors='ignore')
                backup_file = open(backup_file_str, 'r', errors='ignore')

                diff = difflib.context_diff(backup_file.readlines(), orig_file.readlines())
                for l in diff:
                    tempResult = findLineMarker(l)
                    if tempResult is not None:
                        diff_list.append([tempResult.group(2), tempResult.group(3)])

            # RECOVER OLD JUSTIFICATION / ACTION / STATUS
            # Check if old report exists
            if not diff_list:
                if os.path.isfile(old_path):
                    # Check line number AND Warning code AND Source content
                    if (old_sh.cell(row = row, column = 1).value == ws.cell(row = row, column = 1).value and
                        old_sh.cell(row = row, column = 4).value == ws.cell(row = row, column = 4).value and
                        old_sh.cell(row = row, column = 6).value == ws.cell(row = row, column = 6).value) :

                        # copy old justification
                        ws.cell(row = row, column = 7).value = old_sh.cell(row = row, column = 7).value
                        # copy old action
                        ws.cell(row = row, column = 8).value = old_sh.cell(row = row, column = 8).value
                        # copy old status
                        ws.cell(row = row, column = 9).value = old_sh.cell(row = row, column = 9).value
            else:
                new_c_line = int(ws.cell(row=row, column=1).value)

                if section < len(diff_list):
                    if new_c_line < int(diff_list[section+1][0]) or new_c_line > int(diff_list[section+1][1]):
                        if new_c_line >= int(diff_list[section+1][1]):
                            section += 2

                if section == 0:
                    # Check line number AND Warning code AND Source content
                    if (old_sh.cell(row = row, column = 1).value == ws.cell(row = row, column = 1).value and
                        old_sh.cell(row = row, column = 4).value == ws.cell(row = row, column = 4).value and
                        old_sh.cell(row = row, column = 6).value == ws.cell(row = row, column = 6).value) :

                        # copy old justification
                        ws.cell(row = row, column = 7).value = old_sh.cell(row = row, column = 7).value
                        # copy old action
                        ws.cell(row = row, column = 8).value = old_sh.cell(row = row, column = 8).value
                        # copy old status
                        ws.cell(row = row, column = 9).value = old_sh.cell(row = row, column = 9).value
                else:
                    end_offset = int(diff_list[section - 1][1]) - int(diff_list[section - 2][1])
                    # Check line number AND Warning code AND Source content
                    for magic_row_idx in range(5, old_sh.max_row + 1):
                        if (old_sh.cell(row = magic_row_idx, column = 1).value == (ws.cell(row = row, column = 1).value - end_offset) and
                            old_sh.cell(row = magic_row_idx, column = 4).value == ws.cell(row = row, column = 4).value and
                            old_sh.cell(row = magic_row_idx, column = 6).value == ws.cell(row = row, column = 6).value) :

                            # copy old justification
                            ws.cell(row = row, column = 7).value = old_sh.cell(row = magic_row_idx, column = 7).value
                            # copy old action
                            ws.cell(row = row, column = 8).value = old_sh.cell(row = magic_row_idx, column = 8).value
                            # copy old status
                            ws.cell(row = row, column = 9).value = old_sh.cell(row = magic_row_idx, column = 9).value

    wb.save(f)

    if os.path.exists(backup_file_str):
        orig_file.close()
        backup_file.close()

        shutil.copyfile(orig_file_str, backup_file_str)

def extractMetricsReport(fileName):
    global workingFolder
    f_input = os.path.join(workingFolder, fileName + ".met")
    res = []
    with open(f_input, "rt", errors='ignore') as file:
        for line in file:
            if re.match("<S>STFIL.*"+fileName, line.strip()):
                d = dict()
                for line2 in file:
                    m = re.match("<S>(\w+) *(.+)", line2.strip())
                    if m:
                        d[m.group(1)] = m.group(2).strip()
                    else:
                        break
                if 'STNAM' in d:
                    name = d.pop('STNAM')
                    if not name.endswith(fileName):
                        res.append([name, d])
        return {'name': fileName, 'metrics': res, 'FileMetrics': d}


def get_project_metrics(temp_patch):
    project_metrics_list = []
    # keywords = ["STCYA", "STNFA", "STNEA", "STNRA" ]
    f_input = os.path.join(temp_patch, "Output.met")
    with open(f_input, "rt", errors='ignore') as file:
        for line in file:
            if "STCYA" in line:
                m = re.match("(<S>STCYA )(.*)", line)
                project_metrics_list.append(m.group(2))

            if "STNFA" in line:
                m = re.match("(<S>STNFA )(.*)", line)
                project_metrics_list.append(m.group(2))

            if "STNEA" in line:
                m = re.match("(<S>STNEA )(.*)", line)
                project_metrics_list.append(m.group(2))

            if "STNRA" in line:
                m = re.match("(<S>STNRA )(.*)", line)
                project_metrics_list.append(m.group(2))

    return project_metrics_list


def extractCMAMetricsReport(fileName):
    global workingFolder
    f_input = os.path.join("S:\\Metrics\\Static_Analysis\\Temp", fileName + ".met")
    res = []
    with open(f_input, "rt", errors='ignore') as file:
        for line in file:
            if re.match("<S>STFIL.*"+fileName, line.strip()):
                d = dict()
                for line2 in file:
                    m = re.match("<S>(\w+) *(.+)", line2.strip())
                    if m:
                        d[m.group(1)] = m.group(2).strip()
                    else:
                        break
                if 'STNAM' in d:
                    name = d.pop('STNAM')
                    if not name.endswith(fileName):
                        res.append([name, d])
    return {'name' : fileName, 'metrics' : res, 'FileMetrics' : d}

def generateMetricsReport(fullAnalysis):
    global workingFolder
    global swComponentTrigram
    global swProject

    f_output = os.path.join(workingFolder, swComponentTrigram + "_Metrics.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    metricNames = ["STAKI", "STAV1", "STAV2", "STAV3", "STBAK", "STCAL", "STCYC", "STELF", "STFN1", "STFN2", "STGTO", "STKDN", "STKNT", "STLCT", "STLIN", "STLOP", "STM07", "STM19", "STM29",
                   "STMCC", "STMIF", "STPAR", "STPBG", "STPDN", "STPTH", "STRET", "STST1", "STST2", "STST3", "STSUB", "STUNR", "STUNV", "STXLN"]
    headers = ["Function Name", "Define Information", "Define Location"] + metricNames
    essentialMetrics = {
        "STBAK": {'text': "Number of backward jumps", 'max': 0},
        "STCYC": {'text': "KGAS_2770 - Cyclomatic complexity", 'max': 10},
        "STGTO": {'text': "KGAS_3465- Number of Gotos", 'max': 0},
        "STKNT": {'text': "Knot count", 'max': 0},
        "STLIN": {'text': "Number of code lines", 'max': 99},
        "STMIF": {'text': "KGAS_3469 - Deepest level of nesting", 'max': 4},
        "STPTH": {'text': "KGAS_2769 - Estimated static program paths", 'max': 80},
        "STSUB": {'text': "KGAS_3466 - Number of function calls", 'max': 7},
        "STPAR": {'text': "KGAS_3467 - Number of function parameters", 'max': 5},
        "STST3": {'text': "KGAS_3468 - Number of statement in a function", 'max': 50},
        "STM19": {'text': "KGAS_3470 - Number of return statement", 'max': 1},
        "STCAL": {'text': "KGAS_2771 - Number of functions called from function", 'max': 5},
    }

    redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # BACKGROUND COLORS for header cells
    color_cells = "C0C0C0"
    fill_cells = PatternFill("solid", fgColor=color_cells)

    # BORDERS for header cells
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    bottom = Border(bottom=border.bottom)
    right = Border(right=border.right)
    top = Border(top=border.top)
    square = right + bottom
    more = square + top

    improvableFile = []
    improvableFunction = []
    improvableFileVersion = []
    index_ImprovableFunction = 0

    ws_Improvable = wb.create_sheet(0)
    ws_Improvable.title = "Improvable"

    for srcFileAnalyzed in fullAnalysis:
        sheet_name = "Functions Metrics of " + srcFileAnalyzed['name']
        ws = wb.create_sheet(title = sheet_name[0:31])

        ws['A1'] = "Project " + swProject
        ws['A2'] = "Metrics report of " + srcFileAnalyzed['name'] + " " + srcFileAnalyzed['version']
        ws['A3'] = "Generated at " + str(datetime.datetime.now())

        ws['A4'].border = more
        ws['A5'].border = square
        ws['A5'].fill = fill_cells

        # Header
        for idx, string in enumerate(headers):
            ws.cell(row=4,column=idx+1).value = string
            ws.cell(row=4,column=idx+1).border = more
            ws.cell(row=4,column=idx+1).fill = fill_cells

        # For all functions
        func_max_length = 0
        for idx, func in enumerate(srcFileAnalyzed['metrics']):
            func_name, func_metrics = func
            func_max_length = max(func_max_length, len(func_name))
            row = idx + 6
            ws.cell(row=row, column=1).value = func_name
            ws.cell(row=row, column=1).border = square
            metricIssue = False
            for col_idx, metricName in enumerate(metricNames):
                col = col_idx + 4
                if metricName in func_metrics:
                    try:
                        value = int(func_metrics[metricName])
                        if metricName in essentialMetrics and 'max' in essentialMetrics[metricName]:
                            if value > essentialMetrics[metricName]['max']:
                                metricIssue = True
                                ws.cell(row=row, column=col).fill = redFill
                        if metricName == "STBAK":
                            value_STBAK = value
                        if metricName == "STCYC" :
                            value_STCYC = value
                        if metricName == "STGTO":
                            value_STGTO = value
                        if metricName == "STKNT" :
                            value_STKNT = value
                        if metricName == "STLIN" :
                            value_STLIN = value
                        if metricName == "STMIF" :
                            value_STMIF = value
                        if metricName == "STPTH" :
                            value_STPTH = value
                        if metricName == "STSUB":
                            value_STSUB = value
                        if metricName == "STPAR":
                            value_STPAR = value
                        if metricName == "STST3":
                            value_STST3 = value
                        if metricName == "STM19":
                            value_STM19 = value
                        if metricName == "STCAL":
                            value_STCAL = value
                    except ValueError:
                        value = func_metrics[metricName]
                else:
                    value = ""
                ws.cell(row=row, column=col).value = value
                ws.cell(row=row, column=col).border = square

            if metricIssue:
                ws.cell(row=row, column=1).fill = redFill
            try:
                ws.column_dimensions['B'].hidden = True
                ws.column_dimensions['C'].hidden = True
                for col_idx, c_name in enumerate(metricNames):
                    col = col_idx + 4
                    if c_name in essentialMetrics:
                        ws.cell(row=5, column=col).value = essentialMetrics[c_name]['text']
                        ws.cell(row=5, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
                        ws.cell(row=5, column=col).border = square
                        ws.cell(row=5, column=col).fill = fill_cells
                    # else:
                        # col_str = ws.cell(row=5, column=col).column
                        # ws.column_dimensions[col_str].hidden = True
            except:
                print("An exception occurred")

            Rule_1 = (value_STBAK > essentialMetrics["STBAK"]["max"])
            Rule_2 = (value_STCYC > essentialMetrics["STCYC"]["max"])
            Rule_3 = (value_STGTO > essentialMetrics["STGTO"]["max"])
            Rule_4 = (value_STKNT > essentialMetrics["STKNT"]["max"])
            Rule_5 = (value_STLIN > essentialMetrics["STLIN"]["max"])
            Rule_6 = (value_STMIF > essentialMetrics["STMIF"]["max"])
            Rule_7 = (value_STPTH > essentialMetrics["STPTH"]["max"])
            Rule_8 = (value_STSUB > essentialMetrics["STSUB"]["max"])
            Rule_9 = (value_STPAR > essentialMetrics["STPAR"]["max"])
            Rule_10 = (value_STST3 > essentialMetrics["STST3"]["max"])
            Rule_11 = (value_STM19 > essentialMetrics["STM19"]["max"])
            Rule_12 = (value_STCAL > essentialMetrics["STCAL"]["max"])

            if Rule_1 or Rule_2 or Rule_3 or Rule_4 or Rule_5 or Rule_6 or Rule_7 or Rule_8 or Rule_9 or Rule_10 or Rule_11 or Rule_12:
                improvableFunction.append(func_name)
                improvableFile.append(srcFileAnalyzed['name'])
                improvableFileVersion.append(srcFileAnalyzed['version'])

        ws.column_dimensions['A'].width  = func_max_length

        # Sheet for analysis of file based metrics
        FileMetricNames = ["STBME", "STBMO", "STBMS", "STBUG", "STCDN", "STDEV", "STDIF", "STECT", "STEFF", "STFCO", "STFNC", "STHAL", "STM20", "STM21", "STM22", "STM28", "STM33", "STOPN",
                           "STOPT", "STSCT", "STSHN", "STTDE", "STTDO", "STTDS", "STTLN", "STTOT", "STTPP", "STVAR", "STVOL", "STZIP", "VOCF", "COMF"]
        FileMetricsheaders = FileMetricNames
        essentialFileMetrics = {
            "COMF": {'text': "KGAS_3462 - Ratio Nb Comments vs nb of statement", 'min': 0.2},
            "VOCF": {'text': "KGAS_3471 - Vocabulary frequency", 'max': 4}
        }
        sheet_name = "File Metrics of " + srcFileAnalyzed['name']
        ws = wb.create_sheet(title=sheet_name[0:31])

        ws['A1'] = "Project " + swProject
        ws['A2'] = "Metrics report of " + srcFileAnalyzed['name'] + " " + srcFileAnalyzed['version']
        ws['A3'] = "Generated at " + str(datetime.datetime.now())

        ws['A4'].border = more
        ws['A5'].border = square
        ws['A5'].fill = fill_cells

        # Header
        ws.cell(row=4, column=1).value = "Metric name"
        ws.cell(row=4, column=1).border = square
        ws.cell(row=4, column=1).fill = fill_cells

        for idx, string in enumerate(FileMetricsheaders):
            ws.cell(row=4, column=idx + 2).value = string
            ws.cell(row=4, column=idx + 2).border = more
            ws.cell(row=4, column=idx + 2).fill = fill_cells

        row = 6
        ws.cell(row=row, column=1).value = srcFileAnalyzed['name']
        ws.cell(row=row, column=1).border = square
        ws.cell(row=row, column=1).fill = fill_cells
        metricIssue = False
        for col_idx, metricName in enumerate(FileMetricNames):
            col = col_idx + 2
            if metricName in srcFileAnalyzed['FileMetrics']:
                try:
                    value = float(srcFileAnalyzed['FileMetrics'][metricName])
                    if metricName in essentialFileMetrics and 'max' in essentialFileMetrics[metricName]:
                        if value > essentialFileMetrics[metricName]['max']:
                            metricIssue = True
                            ws.cell(row=row, column=col).fill = redFill
                    if metricName in essentialFileMetrics and 'min' in essentialFileMetrics[metricName]:
                        if value < essentialFileMetrics[metricName]['min']:
                            metricIssue = True
                            ws.cell(row=row, column=col).fill = redFill
                    if metricName == "STM20":
                        value_STM20 = value
                    if metricName == "STM21":
                        value_STM21 = value
                    if metricName == "STM22":
                        value_STM22 = value
                    if metricName == "STM28":
                        value_STM28 = value
                    if metricName == "STOPT":
                        value_STOPT = value
                    if metricName == "STOPN":
                        value_STOPN = value

                except ValueError:
                    value = func_metrics[metricName]
            else:
                if metricName == "VOCF":
                    value_VOCF = (value_STM20 + value_STM21) / (value_STOPT + value_STOPN)
                    value = value_VOCF

                    if metricName in essentialFileMetrics and 'max' in essentialFileMetrics[metricName]:
                        if value > essentialFileMetrics[metricName]['max']:
                            metricIssue = True
                            ws.cell(row=row, column=col).fill = redFill

                elif metricName == "COMF":
                    if value_STM22 == 0:
                        value_COMF = 255
                    else:
                        value_COMF = value_STM28 / value_STM22
                    value = value_COMF

                    if metricName in essentialFileMetrics and 'min' in essentialFileMetrics[metricName]:
                        if value < essentialFileMetrics[metricName]['min']:
                            metricIssue = True
                            ws.cell(row=row, column=col).fill = redFill
                else:
                    value = ""
            ws.cell(row=row, column=col).value = value
            ws.cell(row=row, column=col).border = square

            if metricIssue:
                ws.cell(row=row, column=1).fill = redFill

            for col_idx, c_name in enumerate(FileMetricNames):
                col = col_idx + 2
                if c_name in essentialFileMetrics:
                    ws.cell(row=5, column=col).value = essentialFileMetrics[c_name]['text']
                    ws.cell(row=5, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
                    ws.cell(row=5, column=col).border = square
                    ws.cell(row=5, column=col).fill = fill_cells
                # else:
                    # col_str = ws.cell(row=5, column=col).column
                    # ws.column_dimensions[col_str].hidden = True

        # To check criteria of improvements
        Rule_1 = (value_COMF < essentialFileMetrics["COMF"]["min"])
        Rule_2 = (value_VOCF > essentialFileMetrics["VOCF"]["max"])

        if Rule_1 or Rule_2:
            # print("rules out of threshold")
            improvableFile.append(srcFileAnalyzed['name'])
            improvableFileVersion.append(srcFileAnalyzed['version'])

        ws.column_dimensions['A'].width = func_max_length

    # To prepare the "Improvable" sheet
    ws_Improvable['A1'] = "Project " + swProject
    ws_Improvable['A2'] = "Generated at " + str(datetime.datetime.now())

    ws_Improvable.column_dimensions['A'].width  = 60
    ws_Improvable.column_dimensions['B'].width  = 90
    ws_Improvable['B3'] = "The following functions must be improved on next release :"
    ws_Improvable['A3'].fill = fill_cells
    ws_Improvable['B3'].fill = fill_cells
    ws_Improvable['A3'].border = more
    ws_Improvable['B3'].border = more

    for index_ImprovableFunction in range(0,len(improvableFunction)) :
        ws_Improvable.cell(row=4+index_ImprovableFunction, column=1).value = improvableFile[index_ImprovableFunction] + " " + improvableFileVersion[index_ImprovableFunction]
        ws_Improvable.cell(row=4+index_ImprovableFunction, column=2).value = improvableFunction[index_ImprovableFunction]
        ws_Improvable.cell(row=4+index_ImprovableFunction, column=1).border = square
        ws_Improvable.cell(row=4+index_ImprovableFunction, column=2).border = square
    if len(improvableFunction) == 0 :
        ws_Improvable['B4'] = "none"

    # To save the work book
    wb.save(f_output)


def analyzeFile(fileFolder, fileName, htmlReport = True, warningReport = True):
    if not analyzeCFile(fileFolder, fileName):
        print("QAC analysis went wrong")
        return None

    if htmlReport:
        #Generate Html report in workingFolder
        generateHtmlReport(fileFolder, fileName)

    extractWarningsToTxt(fileFolder, fileName) #Generate temporary fileName + "_warning.txt"
    warningList = extractWarningsToList(fileName)

    srcFile = readSrcFile(fileFolder, fileName)
    fileVersion = getPTCFileVersion(srcFile)
    print('Parsing file "' + fileName + '" with version "' + fileVersion + '" !')

    addSrcFileInformationToWarningsList(srcFile, warningList)
    QACJustifications = searchForQACJustification(srcFile)
    addQACJustificationToWarningsList(warningList, QACJustifications)

    searchForPRQA(srcFile, fileName, fileFolder)

    if warningReport:
        generateWarningReport(warningList, fileName, fileVersion) #Generate fileName + "_WarningReport.xlsx"

    metrics = extractMetricsReport(fileName)
    # clearTempFiles(fileName)
    metrics['warnings'] = warningList
    metrics['version'] = fileVersion
    return metrics

def createCMAFiles(folderPath):
    global filesToAnalyzeAVA

    for root, subFolders, files in os.walk(folderPath):
        for file in [file for file in files if file.endswith(".c")]:
            filesToAnalyzeAVA.append([root, file])

def runCMAAnalysis():

    # prepare CMA (Cross Module Analysis) - to get all projet / file metrics
    oldComponent = ""
    commandText = """""";

    # Parse list of analyzed file to create Filelist.lst file
    for index in filesToAnalyzeAVA:

        Output = re.match(r"S:\\Components\\Application\\Autoliv\\(\w+)", str(index[0]))
        if Output is not None:
            component = Output.group(1)

            if component != oldComponent:
                analysisfolder = "S:\\Components\\Application\\Autoliv\\" + component + "\\Quality_Assurance\\Static_Analysis"
                commandText += """-op \"""" + analysisfolder + """\"\n"""
                oldComponent = component

            ComponentFolder = "S:\\Components\\Application\\Autoliv\\" + component + "\\Implementation\\src\\"
            commandText += """\"""" + ComponentFolder + str(index[1]) + """\"\n"""

    FileOut = open("S:\\Metrics\\Static_Analysis\\Temp\\Filelist.lst", "w", errors='ignore')
    FileOut.write(commandText)
    FileOut.close()

    if not os.path.exists("S:\\Metrics\\Static_Analysis\\Temp"):
        os.makedirs("S:\\Metrics\\Static_Analysis\\Temp")

    # Create command for CMA
    command = [os.path.join(QAC_folder, "bin", "pal.exe"), "QAC"]
    command += ["-cmaf", "S:\\Metrics\\Static_Analysis\\Temp\\Output"]
    command += ["-list", "S:\\Metrics\\Static_Analysis\\Temp\\Filelist.lst"]

    # perform the command
    subprocess.check_call(command)

def analyzeFolder(folderPath, htmlReport = True, warningReport = True):
    filesToAnalyze = []
    global listToBeAnalysis
    global fileListToBeClear

    for root, subFolders, files in os.walk(folderPath):
        for file in [file for file in files if file.endswith(".c")]:
            filesToAnalyze.append([root, file])
    analysis = []
    for folder, file in filesToAnalyze:
        print("Analysing " + file)
        analysis.append(analyzeFile(folder, file, htmlReport = htmlReport, warningReport = warningReport))
        fileListToBeClear.append(file)

    listToBeAnalysis.append(analysis)

    # if metricsReport:
        # generateMetricsReport(analysis)
    return analysis



def moveFileToComponentLibrary(componentLibraryPath):
    global workingFolder
    fileList = os.listdir(workingFolder)
    files = [f for f in fileList if (f.endswith(".xlsx") or f.endswith(".html"))]
    for f in files:
        in_path = os.path.join(workingFolder, f)
        out_path = os.path.join(componentLibraryPath, f)
        os.rename(in_path, out_path)

def generateComponentReport(folderPath):
    global workingFolder
    global swProject
    global swComponentTrigram

    metricsFile = os.path.join(workingFolder, swComponentTrigram + "_Metrics.xlsx")
    metrixFile = os.path.join(workingFolder, swComponentTrigram + "_Metrix.xlsx")
    componentFileName = swComponentTrigram + "_QACReport.xlsx"
    componentFilePath = os.path.join(workingFolder, componentFileName)

    # Create the Component report
    if os.path.isfile(metricsFile):
        shutil.copy(metricsFile, componentFilePath)
    if os.path.isfile(metrixFile):
        shutil.copy(metrixFile, componentFilePath)

def return_list_with_file_names(modules_path):
    def return_module_names(module_path):
        module_names = []
        module_paths = glob.glob(module_path + '\*')
        for x in os.listdir(module_path):
            module_names.append(x)
        return module_names, module_paths

    def find_files(pattern, path):
        result = []
        for root, dirs, files in os.walk(path):
            for name in files:
                if fnmatch.fnmatch(name, pattern):
                    result.append(os.path.join(root, name))
        return result

    def return_list_of_file_paths(module_paths):
        list_with_file_paths = []
        for i in module_paths:
            list_with_file_paths.append(find_files('*_Metrics.xlsx', i))
        return list_with_file_paths
    module_names, module_paths = return_module_names(module_path=modules_path)
    list_with_file_names = return_list_of_file_paths(module_paths=module_paths)
    return list_with_file_names


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def get_data(metrics_file_name):

    list_with_sheet_names = []
    final_list_with_sheet_names = []

    improvable_data_to_write = []
    improvable_data_to_write_final = []


    metrics_list = []
    metrics_file_list = []

    for metrics_files in metrics_file_name:
        if metrics_files != []:
            wb = openpyxl.load_workbook(metrics_files[0])
            list_with_sheet_names.append(wb.sheetnames )
            for sheet in wb.worksheets:
                if "Improvable" in sheet.title:
                    max = sheet.max_row
                    argument = 'B' + str(max)
                    for row in sheet['A4':argument]:
                        row_data_from_improvable = []
                        for cell in row:
                            if cell.value != 'none' and cell.value is not None:
                                row_data_from_improvable.append(cell.value)
                        improvable_data_to_write.append(row_data_from_improvable)

                elif "Functions Metrics" in sheet.title:
                    max_c = sheet.max_column
                    max_r = sheet.max_row
                    argument2 = str(colnum_string(max_c)) + str(max_r)
                    metrics_sheets_list = []
                    for row in sheet['A6':argument2]:
                        row_data_for_other_metrics = []
                        for cell in row:
                            if cell.value is None:
                                cell.value = ''
                            row_data_for_other_metrics.append(cell.value)
                        metrics_sheets_list.append(row_data_for_other_metrics)
                    metrics_list.append(metrics_sheets_list)
                elif "File Metrics" in sheet.title:
                    max_c = sheet.max_column
                    max_r = sheet.max_row
                    argument2 = str(colnum_string(max_c)) + str(max_r)
                    file_metrics_sheets_list = []
                    for row in sheet['A6':argument2]:
                        row_data_for_other_metrics = []
                        for cell in row:
                            if cell.value is None:
                                cell.value = ''
                            row_data_for_other_metrics.append(cell.value)
                        file_metrics_sheets_list.append(row_data_for_other_metrics)
                    metrics_file_list.append(file_metrics_sheets_list)
    #Make some clean in Improvable list
    for elements in improvable_data_to_write:
        if elements != []:
            improvable_data_to_write_final.append(elements)

    for semi in list_with_sheet_names:
        for i in range(1, len(semi)): #1 ca sa ignore primul element adica Improvable si 0 ca sa afiseze si improvable
            if "Functions Metrics" in semi[i]:
                final_list_with_sheet_names.append(semi[i][21:])

    return final_list_with_sheet_names,metrics_list,improvable_data_to_write_final, metrics_file_list


def write_in_excel(final_list_with_sheet_names,metrics_list,improvable_data_to_write_final,metrics_file_list, input_metrics_template, output_generated_file, project_metrics_list):
    wb = openpyxl.load_workbook(input_metrics_template)
    sheet = wb['Improvable']
    for row in range(0, len(improvable_data_to_write_final)):
        for col in range(0, len(improvable_data_to_write_final[row])):
            sheet.cell(row = 4 + row , column = 1 + col ).value = improvable_data_to_write_final[row][col]

    i = 1
    j = 1
    index_for_names = 0

    offset_test = 0
    sheet2 = wb['Functions Metrics']
    for sheet_lists in range(0,len(metrics_list)):
        sheet2.cell(row = 6 + offset_test, column = 1).value = final_list_with_sheet_names[index_for_names]
        for rows_data in metrics_list[sheet_lists]:
            j = 1
            for columns_data in rows_data:
                sheet2.cell(row= i + 5 , column= j + 1).value = columns_data
                j += 1
            i += 1
        index_for_names += 1
        offset_test = offset_test + len(metrics_list[sheet_lists])
    i = 1
    j = 1
    index_for_names = 0

    offset_test = 0
    sheet2 = wb['File Metrics']
    for sheet_lists in range(0,len(metrics_file_list)):
        # sheet2.cell(row = 6 + offset_test, column = 1).value = final_list_with_sheet_names[index_for_names]
        for rows_data in metrics_file_list[sheet_lists]:
            j = 0
            for columns_data in rows_data:
                sheet2.cell(row= i + 5 , column= j + 1).value = columns_data
                j += 1
            i += 1
        index_for_names += 1
        offset_test = offset_test + len(metrics_file_list[sheet_lists])
    i = 1
    j = 1
    sheet2 = wb['Project Metrics']
    for date in project_metrics_list:
        sheet2.cell(row=i + 5, column=j + 1).value = date
        j += 1

    wb.save(output_generated_file)
    wb.close()


def get_statics(input_components_path):
    excel_output_list = []

    currentDir = os.getcwd()
    sys.path.append(currentDir)

    warnings.simplefilter('ignore', UserWarning)

    def path_leaf(path):
        head, tail = ntpath.split(path)
        return tail or ntpath.basename(head)

    application_path = [input_components_path]

    project_statistics = {}

    for path in application_path:
        component_list = os.listdir(path)
        # csv_output_str = os.path.normpath(os.path.join(currentDir, "{}.csv".format('all')))
        # print(csv_output_str)
        for component in component_list:

            # Initialize dictionary that holds all statistics
            if component not in project_statistics:
                project_statistics[component] = {}

            search_list = glob.iglob(
                '{}/{}/Quality_Assurance/Static_Analysis/**/*.c_WarningReport.xlsx'.format(path, component),
                recursive=True)

            for static_analysis_file in search_list:
                if path_leaf(static_analysis_file) not in project_statistics[component]:
                    project_statistics[component][path_leaf(static_analysis_file)] = [
                        [],  # Warning Type
                        [],  # Severity
                        [],  # Status
                        []  # Action
                    ]

                print("Parsing {}".format(path_leaf(static_analysis_file)))

                wb = load_workbook(filename=static_analysis_file, data_only=True)

                # Get "Code" sheet
                curr_sheet = wb.active

                # Read Error Type
                for col in curr_sheet.iter_rows(min_row=5, min_col=2, max_col=2):
                    for cell in col:
                        project_statistics[component][path_leaf(static_analysis_file)][0].append(cell.value)

                # Read Severity
                for col in curr_sheet.iter_rows(min_row=5, min_col=3, max_col=3):
                    for cell in col:
                        project_statistics[component][path_leaf(static_analysis_file)][1].append(cell.value)

                # Read Status
                for col in curr_sheet.iter_rows(min_row=5, min_col=9, max_col=9):
                    for cell in col:
                        project_statistics[component][path_leaf(static_analysis_file)][2].append(cell.value)

                # Read Action
                for col in curr_sheet.iter_rows(min_row=5, min_col=8, max_col=8):
                    for cell in col:
                        project_statistics[component][path_leaf(static_analysis_file)][3].append(cell.value)

    for module, all_stats in project_statistics.items():
        qac_warnings = 0
        severe_low_qac_warnings = 0
        severe_medium_qac_warnings = 0
        severe_high_qac_warnings = 0
        warnings_to_be_corrected = 0
        for file, stats in all_stats.items():
            qac_warnings += len(stats[0])

            for idx in range(0, len(stats[1])):
                try:
                    if stats[1][idx] is not None:

                        if stats[1][idx] <= 2:
                            severe_low_qac_warnings += 1
                        elif stats[1][idx] <= 5:
                            severe_medium_qac_warnings += 1
                        elif stats[1][idx] > 5:
                            severe_high_qac_warnings += 1

                except:
                    pass

            for idx in range(0, len(stats[2])):
                try:
                    if ((stats[2][idx] is None) and (stats[0][idx] is not None )):
                        warnings_to_be_corrected += 1
                except:
                    pass
                    
        excel_output_list.append([
            module,
            qac_warnings,
            severe_low_qac_warnings,
            severe_medium_qac_warnings,
            severe_high_qac_warnings,
            warnings_to_be_corrected
        ])
    return excel_output_list


def insert_rows(number_of_lines, input_template_path, output_generated_file_path):
    xl = client.Dispatch("Excel.Application")
    wb = xl.Workbooks.Open(input_template_path)
    xl.Visible = 1
    ws = wb.Sheets("metrics")
    start_row_to_insert = 9
    for i in range(1, number_of_lines):
        x = ws.Range('A9:W10')
        x.EntireRow.Insert()
    for i in range(1, number_of_lines):
        ws.Range('A7:W8').Copy()
        insert_to_parameter = 'A' + str(start_row_to_insert)
        start_row_to_insert += 2
        ws.Paste(ws.Range(insert_to_parameter))
    xl.ActiveWorkbook.SaveAs(output_generated_file_path)
    xl.Quit()


def write_in_excel_qac(path_to_file, list_to_write):
    book = openpyxl.load_workbook(path_to_file)
    sheet = book.active
    index_for_list_to_write = 0
    index_for_elements_in_every_row = 0

    for row in sheet.iter_rows(min_row=7, min_col=1, max_col=6):
        for cell in row:
            cell.value = list_to_write[index_for_list_to_write][index_for_elements_in_every_row]
            if index_for_elements_in_every_row == 5:
                index_for_elements_in_every_row = 0
            else:
                index_for_elements_in_every_row += 1
        index_for_list_to_write += 1
        if (index_for_list_to_write == len(list_to_write)):
            break
    book.save(path_to_file)
    book.close()


# Select QAC intallation folder
QAC_folder = "C:/PRQA/QAC-8.1.2-R"
os.environ["QACBIN"] = os.path.join(QAC_folder, "bin")
os.environ["QACHELPFILES"] = os.path.join(QAC_folder, "help")
os.environ["QACOUTPATH"] = os.path.join(QAC_folder, "temp")

# Select the PTC project, which will be mentionned in the QAC generated reports
swProject = "SBE/PP/DAIMLER_MMA/"

# Select QAC tool folder
f = "S:\\Metrics\\Static_Analysis\\"
configFiles = []
configFiles.append(os.path.join(f, "AUTOLIV.p_c"))
configFiles.append(os.path.join(f, "AUTOLIV.p_s"))
configFiles.append(os.path.join(f, "PP4G_ERR.p_a"))

patch_dict = {
    "input_components_path"         : "S:\\Components\\Application\\Autoliv",
    "input_template_path_metrics"   : "S:\\Metrics\\Static_Analysis\\Template\\QAC_Metrics_Report_Template.xlsx",
    "output_generated_path_metrics" : "S:\\Metrics\\Static_Analysis\\QAC_Metrics_Report.xlsx",
    "input_template_path_qac"       : "S:\\Metrics\\Static_Analysis\\Template\\QAC_Warning_Report_Template.xlsx",
    "output_generated_path_qac"     : "S:\\Metrics\\Static_Analysis\\QAC_Warning_Report.xlsx",
    "temp_patch"                    : "S:\\Metrics\\Static_Analysis\\Temp"}

parser = argparse.ArgumentParser()
parser.add_argument('-c', '--component', help="Name of the component", required=False, default="all")
args = parser.parse_args()

# Select SWC to analyse
array_swComponentTrigram = [
     "ERH",
]


# array_swComponentTrigram = os.listdir(patch_dict["input_components_path"])

swComponentTrigram = "all"
# swComponentTrigram = "AdcIf"
# swComponentTrigram = "ATM"
# swComponentTrigram = "BFD"
# swComponentTrigram = "BFE"
# swComponentTrigram = "BFS"
# swComponentTrigram = "BMM"
# swComponentTrigram = "BPA"
# swComponentTrigram = "BSR"
# swComponentTrigram = "BswMIf"
# swComponentTrigram = "CanTrcvIf"
# swComponentTrigram = "CIL"
# swComponentTrigram = "CompilerIf"
# swComponentTrigram = "DcmIf"
# swComponentTrigram = "DemIf"
# swComponentTrigram = "DIA"
# swComponentTrigram = "DioIf"
# swComponentTrigram = "ECUMIf"
# swComponentTrigram = "EFX"
# swComponentTrigram = "EOL"
# swComponentTrigram = "FlsIf"
# swComponentTrigram = "FSM"
# swComponentTrigram = "HWA"
# swComponentTrigram = "MIC"
# swComponentTrigram = "MMG"
# swComponentTrigram = "NmIf"
# swComponentTrigram = "NvMIf"
# swComponentTrigram = "NVP"
# swComponentTrigram = "OsIf"
# swComponentTrigram = "PAL"
# swComponentTrigram = "PCM"
# swComponentTrigram = "PlatformsIf"
# swComponentTrigram = "PMP"
# swComponentTrigram = "PortIf"
# swComponentTrigram = "PRE"
# swComponentTrigram = "PRO"
# swComponentTrigram = "PwmIf"
# swComponentTrigram = "RCM"
# swComponentTrigram = "RTE"
# swComponentTrigram = "RteIf"
# swComponentTrigram = "SBC"
# swComponentTrigram = "SCM"
# swComponentTrigram = "SFR"
# swComponentTrigram = "SpiIf"
# swComponentTrigram = "Startup"
# swComponentTrigram = "STM"
# swComponentTrigram = "TL_Lib"
# swComponentTrigram = "VDA"
# swComponentTrigram = "VIP"


global filesToAnalyzeAVA
global analysis

filesToAnalyzeAVA = []

if args.component == "all" and swComponentTrigram == "all":

    number_of_components = len(array_swComponentTrigram)

    listToBeAnalysis = []
    fileListToBeClear = []
    listWorkingFolder =[];
    # Clear Temp Folder
    for file in os.listdir(patch_dict["temp_patch"]):
        file_path = os.path.join(patch_dict["temp_patch"], file)
        os.remove(file_path)

    for index_component in range(0, number_of_components):

        swComponentTrigram = array_swComponentTrigram[index_component]
        swComponentToAnalyse = "S:\\Components\\Application\\Autoliv\\" + swComponentTrigram + "\\Implementation\\src"
        workingFolder = "S:\\Components\\Application\\Autoliv\\" + swComponentTrigram + "\\Quality_Assurance\\Static_Analysis"
        listWorkingFolder.append(workingFolder);
        # Prepare QAC environment
        createViaFile(configFiles)

        # Launch analysis
        analyzeFolder(swComponentToAnalyse, htmlReport=True, warningReport=True)

    for index_component in range(0, number_of_components):
        swComponentTrigram = array_swComponentTrigram[index_component]
        swComponentToAnalyse = "S:\\Components\\Application\\Autoliv\\" + swComponentTrigram + "\\Implementation\\src"

        # Prepare CMA files
        createCMAFiles(swComponentToAnalyse)

    # Run CMA analysis
    runCMAAnalysis()

    # Generate Metrics reports
    for index_component in range(0, number_of_components):

        swComponentTrigram = array_swComponentTrigram[index_component]
        workingFolder = "S:\\Components\\Application\\Autoliv\\" + swComponentTrigram + "\\Quality_Assurance\\Static_Analysis"
        generateMetricsReport(listToBeAnalysis[index_component])

    # QAC_Metrics_Report
    metrics_file_names = return_list_with_file_names(modules_path=patch_dict["input_components_path"])
    final_list_with_sheet_names, metrics_list, improvable_data_to_write_final, metrics_file_list = get_data(metrics_file_names)
    project_metrics_list=get_project_metrics(patch_dict["temp_patch"])
    if os.path.exists(patch_dict["output_generated_path_metrics"]):
        os.remove(patch_dict["output_generated_path_metrics"])
    write_in_excel(final_list_with_sheet_names, metrics_list, improvable_data_to_write_final, metrics_file_list,
                   patch_dict["input_template_path_metrics"], patch_dict["output_generated_path_metrics"], project_metrics_list)

    # QAC_Warning_Report
    data_list = get_statics(patch_dict["input_components_path"])
    if os.path.exists(patch_dict["output_generated_path_qac"]):
        os.remove(patch_dict["output_generated_path_qac"])
    insert_rows(number_of_lines=round(len(data_list) / 2), input_template_path=patch_dict["input_template_path_qac"],
                output_generated_file_path=patch_dict["output_generated_path_qac"])
    write_in_excel_qac(path_to_file=patch_dict["output_generated_path_qac"], list_to_write=data_list)

     # Clear Temp Files
    for patchWorkingFolder in listWorkingFolder:
        workingFolder = patchWorkingFolder
        for file in fileListToBeClear:
            clearTempFiles(file)
    # Clear Temp Folder
    for file in os.listdir(patch_dict["temp_patch"]):
        file_path = os.path.join(patch_dict["temp_patch"], file)
        os.remove(file_path)

else:
    if swComponentTrigram != "all":
        tempComponent = swComponentTrigram
    else:
        tempComponent = args.component

    swComponentTrigram = tempComponent
    swComponentToAnalyse_src = "S:\\Components\\Application\\Autoliv\\" + swComponentTrigram + "\\Implementation\\src"
    swComponentToAnalyse_cfg = "S:\\Components\\Application\\Autoliv\\" + swComponentTrigram + "\\Implementation\\cfg"
    workingFolder = "S:\\Components\\Application\\Autoliv\\" + swComponentTrigram + "\\Quality_Assurance\\Static_Analysis"
    if not os.path.exists(workingFolder):
        os.makedirs(workingFolder)
    listWorkingFolder = []
    listToBeAnalysis = []
    fileListToBeClear = []
    listWorkingFolder.append(workingFolder)

    # Prepare QAC environment
    createViaFile(configFiles)

    # Launch analysis
    analyzeFolder(swComponentToAnalyse_src, htmlReport=True, warningReport=True)
    analyzeFolder(swComponentToAnalyse_cfg, htmlReport=True, warningReport=True)

    # Prepare CMA files
    createCMAFiles(swComponentToAnalyse_src)
    createCMAFiles(swComponentToAnalyse_cfg)

    # Run CMA analysis
    runCMAAnalysis()

    # Generate Metrics reports
    generateMetricsReport(listToBeAnalysis[0])
