import argparse
import csv
import fnmatch
import glob
import os
from PIL import Image
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

#--------------------------------------------------------
#Updated by Lucian Ardeleanu, 05.10.2020
#Changed plots titles
#---------------------------------------------------------


def return_module_names(module_path):
    module_names = []
    module_paths = glob.glob(module_path + '\*')
    for x in os.listdir(module_path):
        module_names.append(x)
    return module_names, module_paths

def find_files(pattern, path):
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, pattern):
                result.append(os.path.join(root, name))
    return result

def return_list_of_file_paths(module_paths):
    list_with_file_paths = []
    for i in module_paths:
        list_with_file_paths.append(find_files('*.c_WarningReport.xlsx', i))
    return list_with_file_paths

def get_lists_of_warnings(list_with_file_paths):
    list_of_total_warnings = []
    list_of_unjustified_warnings = []
    list_of_severity_warnings = []

    none  = []
    number_of_unjustified_warnings = 0
    number_of_severity_warnings = 0
    number_of_total_warnings = 0
    for module in range(0, len(list_with_file_paths)):
        if list_with_file_paths[module] != none:
            for file in range(0,len(list_with_file_paths[module])):
                    book = openpyxl.load_workbook(list_with_file_paths[module][file])
                    sheet = book.active
                    rows = sheet.max_row
                    max_rows = 1
                    for i in range(1,rows):
                        if ( (sheet.cell(row= i,column=1).value) is not None):
                            max_rows += 1
                    for row in range(5,max_rows):

                        if( (sheet.cell(row= row,column=2).value) is not None):
                            number_of_total_warnings += 1

                        if( (sheet.cell(row= row,column=7).value) is None):
                            number_of_unjustified_warnings += 1

                        if( (sheet.cell(row= row,column=3).value) >= 7 ):
                            number_of_severity_warnings += 1

            list_of_unjustified_warnings.append(number_of_unjustified_warnings)
            list_of_severity_warnings.append(number_of_severity_warnings)
            list_of_total_warnings.append(number_of_total_warnings)



            number_of_severity_warnings = 0
            number_of_unjustified_warnings = 0
            number_of_total_warnings = 0
        else:
            list_of_severity_warnings.append(0)
            list_of_unjustified_warnings.append(0)
            list_of_total_warnings.append(0)

    return list_of_severity_warnings, list_of_unjustified_warnings, list_of_total_warnings

def plot_warning_stats(input_csv_path, output_plot_path, plot_title):
    import csv
    from PIL import Image
    import matplotlib.pyplot as plt
    import numpy as np

#Initialise elements and lists used
    warning_value = []
    build_number = []
    date = []
    module_name = []
    build_number_and_date = []
    enumerate_rows = 0

    data_from_rows = []

#Open .csv file and append all data in a big list
    with open(input_csv_path, 'rt')as f:
        data = csv.reader(f)
        next(csv.reader(f))
        for row in data:
            enumerate_rows += 1
            data_from_rows.append(row)

#Replace empty elements with blank spaces
    for i in range(0, len(data_from_rows)):
        if len(data_from_rows[i]) == 1:
            data_from_rows[i].append(' ')
            data_from_rows[i].append(' ')

#Count in variable: enumerate_rows the last 100 elements used to plot ( we need to plot only last 100 elements from all global warning stats list )
    if len(data_from_rows) > 100:
        enumerate_rows = len(data_from_rows) - 100
    else:
        enumerate_rows = 0


#List is sliced in multiple lists used to plot

    if (len(data_from_rows[1]) == 2):
        for i in range(0, len(data_from_rows)):
            module_name.append(data_from_rows[i][0])
            warning_value.append(data_from_rows[i][1])
    else:
        for i in range(enumerate_rows, len(data_from_rows)):
            warning_value.append(data_from_rows[i][0])
            build_number.append(data_from_rows[i][1])
            date.append(data_from_rows[i][2])
            build_number_and_date.append([])

    if (len(data_from_rows[1]) != 2):
        #Conversions are made for plotting
        for i in range(0, len(warning_value)):
            warning_value[i] = int(warning_value[i])

        for i in range(0, len(build_number)):
            build_number[i] = str(build_number[i])

        for i in range(0, len(date)):
            date[i] = str(date[i])

        for i in range(0, len(build_number)):
            build_number_and_date[i] = '#' + build_number[i] + ' ( ' + date[i] + ' ) '
    else:
        for i in range(0, len(module_name)):
            module_name[i] = str(module_name[i])

        for i in range(0, len(warning_value)):
            warning_value[i] = int(warning_value[i])



#Plot is created using data processed up
    bars = warning_value
    height = warning_value

    y_pos = np.arange(len(bars))
    plt.figure(figsize=(20, 10))
    plt.bar(y_pos, height)

    if (len(data_from_rows[1]) != 2):
        names = build_number_and_date
    else:
        names = module_name
    plt.xticks(y_pos, names, rotation=90)

    # Thus we have to give more margin:
    plt.subplots_adjust(bottom=0.4)

    # It's the same concept if you need more space for your titles
    plt.title(plot_title)
    plt.subplots_adjust(top=0.7)
    plt.savefig(output_plot_path)

    im = Image.open(output_plot_path)

    crop_rectangle = (200, 270, 1810, 800)  # left, top, right, bottom
    cropped_im = im.crop(crop_rectangle)

    newsize = (1450, 400)
    imresize = cropped_im.resize(newsize)

    im = imresize.save(output_plot_path)

def write_in_files(module_names, list_of_severity_warnings, list_of_unjustified_warnings, list_of_total_warnings, output_path, build_number, data):
    module_warning_stats_header = ['Module','No. Of Warnings']

    sum_for_severity_global = 0
    sum_for_unjustified_global = 0
    sum_for_total_global = 0


    lists_for_severity_warnings = zip(module_names,list_of_severity_warnings)
    with open(output_path + '\module_with_severity_warning_stats.csv', "w", newline='') as file:
        file_writer = csv.writer(file )
        file_writer.writerow(module_warning_stats_header)
        file_writer.writerows(lists_for_severity_warnings)
        file.close()

    lists_for_total_warnings = zip(module_names,list_of_total_warnings)
    with open(output_path + '\module_with_total_warning_stats.csv', "w", newline='') as file:
        file_writer = csv.writer(file )
        file_writer.writerow(module_warning_stats_header)
        file_writer.writerows(lists_for_total_warnings)
        file.close()

    lists_for_unjustified_warnings = zip(module_names,list_of_unjustified_warnings)
    with open(output_path + '\module_with_unjustified_warning_stats.csv', "w", newline='') as file:
        file_writer = csv.writer(file )
        file_writer.writerow(module_warning_stats_header)
        file_writer.writerows(lists_for_unjustified_warnings)
        file.close()


    if not os.path.isfile(os.path.normpath(os.path.join(output_path, 'global_with_severity_warning_stats.csv'))):
        global_file_obj = open(os.path.normpath(os.path.join(output_path, 'global_with_severity_warning_stats.csv')), 'w+',  newline='')
        global_file_obj.write('No. Of Warnings\n')
    else:
        global_file_obj = open(os.path.normpath(os.path.join(output_path, 'global_with_severity_warning_stats.csv')), 'a+',  newline='')

    for element in range(0,len(list_of_severity_warnings)):
        sum_for_severity_global = sum_for_severity_global + list_of_severity_warnings[element]

    global_file_obj.write('{},{},{}\n'.format(str(sum_for_severity_global), build_number, data ))
    global_file_obj.close()

    if not os.path.isfile(os.path.normpath(os.path.join(output_path, 'global_with_unjustified_warning_stats.csv'))):
        global_file_obj = open(os.path.normpath(os.path.join(output_path, 'global_with_unjustified_warning_stats.csv')), 'w+',  newline='')
        global_file_obj.write('No. Of Warnings\n')
    else:
        global_file_obj = open(os.path.normpath(os.path.join(output_path, 'global_with_unjustified_warning_stats.csv')), 'a+',  newline='')

    for element in range(0,len(list_of_unjustified_warnings)):
        sum_for_unjustified_global = sum_for_unjustified_global + list_of_unjustified_warnings[element]

    global_file_obj.write('{},{},{}\n'.format(str(sum_for_unjustified_global), build_number , data ))
    global_file_obj.close()

    if not os.path.isfile(os.path.normpath(os.path.join(output_path, 'global_with_total_warning_stats.csv'))):
        global_file_obj = open(os.path.normpath(os.path.join(output_path, 'global_with_total_warning_stats.csv')), 'w+',  newline='')
        global_file_obj.write('No. Of Warnings\n')
    else:
        global_file_obj = open(os.path.normpath(os.path.join(output_path, 'global_with_total_warning_stats.csv')), 'a+',  newline='')

    for element in range(0,len(list_of_total_warnings)):
        sum_for_total_global = sum_for_total_global + list_of_total_warnings[element]

    global_file_obj.write('{},{},{}\n'.format(str(sum_for_total_global), build_number , data ))
    global_file_obj.close()

def main():

    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input-path', help="Path to the components directory, E.G. C:\Application", required=True)
    parser.add_argument('-o', '--output-path', help="Path to the output files, E.G. C:\Output_folder", required=True)
    parser.add_argument('-d', '--data', help="Data to fill in global_warning_stats.csv", required=False)
    parser.add_argument('-b', '--build-number', help="Build number to fill in global_warning_stats.csv", required=False)
    args = parser.parse_args()

    module_names, module_paths = return_module_names(module_path=args.input_path)
    list_of_file_paths = return_list_of_file_paths(module_paths=module_paths)
    list_of_severity_warnings, list_of_unjustified_warnings, list_of_total_warnings = get_lists_of_warnings(list_of_file_paths)
    write_in_files(module_names = module_names, list_of_severity_warnings=list_of_severity_warnings,list_of_unjustified_warnings=list_of_unjustified_warnings,list_of_total_warnings= list_of_total_warnings, output_path=args.output_path, build_number= args.build_number, data = args.data )

    plot_warning_stats(input_csv_path=args.output_path + '\global_with_severity_warning_stats.csv', output_plot_path=args.output_path + '\QAC_Warnings_With_Severity_per_Build_Number.png', plot_title= 'QAC Warnings With Severity >=7 / Build Number')
    plot_warning_stats(input_csv_path=args.output_path + '\global_with_total_warning_stats.csv', output_plot_path=args.output_path + '\QAC_Warnings_per_Build_Number.png', plot_title= 'QAC Warnings / Build Number')
    plot_warning_stats(input_csv_path=args.output_path + '\global_with_unjustified_warning_stats.csv', output_plot_path=args.output_path + r'\Unjustified_QAC_Warnings_per_Build_Number.png', plot_title= 'Unjustified QAC Warnings / Build Number')
    plot_warning_stats(input_csv_path=args.output_path + '\module_with_severity_warning_stats.csv', output_plot_path=args.output_path + '\QAC_Warnings_With_Severity_per_Module.png', plot_title= 'QAC Warnings With Severity >=7 / Module')
    plot_warning_stats(input_csv_path=args.output_path + '\module_with_total_warning_stats.csv', output_plot_path=args.output_path + '\QAC_Warnings_per_Module.png', plot_title= 'QAC Warnings / Module')
    plot_warning_stats(input_csv_path=args.output_path + '\module_with_unjustified_warning_stats.csv', output_plot_path=args.output_path + r'\Unjustified_QAC_Warnings_per_Module.png', plot_title= 'Unjustified QAC Warnings / Module')

if __name__ == "__main__":
    main()