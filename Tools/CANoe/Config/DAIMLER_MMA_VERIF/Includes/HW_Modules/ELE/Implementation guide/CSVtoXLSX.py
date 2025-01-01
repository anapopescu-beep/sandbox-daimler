# install Pillow, OpenPyxl, pyinstaller, zlib
# Package as .exe: pyinstaller -F CSVtoXLSX.py
# .exe needs 4 inputs for .XLSX generation
# input 1 = template.xlsx
# input 2 = result .csv file
# input 3 = template.xlsx sheet to copy .csv data to
# input 4 = output .xlsx file
#
# For zip file creation and deletion:
# input 1 = 'zip'
# input 2 = path to source folder
# input 3 = path and name of resulting archive
import csv
from openpyxl import load_workbook
from copy import copy
import sys
import time
import zipfile
import pathlib
import os
from zipfile import ZipFile, ZIP_DEFLATED
 
if __name__ == "__main__":
    if len(sys.argv)<1:
        raise ValueError("Incorrect input.")

if sys.argv[1] == 'zip':
    directory = pathlib.Path(sys.argv[2])

    with zipfile.ZipFile(sys.argv[3], "w", ZIP_DEFLATED, compresslevel=9) as archive:
        for file_path in directory.iterdir():
            if file_path.suffix == '.asc' or file_path.suffix == '.blf':
                archive.write(file_path, arcname=file_path.name)
                os.remove(file_path)
else:
    workbook = load_workbook(filename = sys.argv[1])
    worksheet = workbook[sys.argv[3]]

    # copy .CSV to xlsx
    with open(sys.argv[2]) as CSVReader:
        ReadCSV = csv.reader(CSVReader, delimiter=";")
        for row in ReadCSV:
            worksheet.append(row)

    # close .CSV
    CSVReader.close()

    # format .xlsx output file
    max_row = worksheet.max_row
    max_col_number = worksheet.max_column
    for row in range(3, max_row + 1):
        for column in range(1, max_col_number + 1):
            if worksheet.cell(row, column).value:
                worksheet.cell(row, column).value = str(str(worksheet.cell(row, column).value).replace("<p>", "\n"))
            worksheet.cell(row, column).fill = copy(worksheet.cell(2, column).fill)
            worksheet.cell(row, column).border = copy(worksheet.cell(2, column).border)

    # close .xlsx
    worksheet.delete_rows(2, 1)
    workbook.save(sys.argv[4])
    workbook.close()