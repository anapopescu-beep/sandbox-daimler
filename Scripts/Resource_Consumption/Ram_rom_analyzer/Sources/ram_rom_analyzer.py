import os
import openpyxl
from openpyxl.styles import *
import argparse
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import yaml
# Function used to convert a string separated with blank spaces to list
#( basically split where are blank spaces and append elements in list )
def convert_strings_to_int(list_to_convert):
    filtered_list = []
    for first_series in list_to_convert:
        second_series = first_series.replace(' ', '')
        if len(second_series) == 0:
            second_series = '0'
        second_series = int(second_series)
        filtered_list.append(second_series)
    return filtered_list


# Function used to convert one element into list
def convert_elements_to_lists_elements(lists):
    return list(map(lambda el:[el], lists))

# Function to process fbl.map and calculate grand rom consumption and grand ram consumption
def process_fbl_map(path_to_fbl_map):

    consumptiondict = {}

    if (path_to_fbl_map != None):
        file0 = open(path_to_fbl_map, 'r')
        Lines0 = file0.readlines()


        for line in Lines0:
            if r'Grand Total:' in line.strip():
                last_line_fbl = line
                break

        # Extract from grand total line ro code, rw code, ro data
        grand_total_rom_consumption_fbl_map = int(last_line_fbl[int(yamlConfig["fbl_rom_ro_code_start"]):int(yamlConfig["fbl_rom_ro_code_end"])].replace(" ", "")) + int(
            last_line_fbl[int(yamlConfig["fbl_rom_ro_data_start"]):int(yamlConfig["fbl_rom_ro_data_end"])].replace(" ", ""))
        grand_total_ram_consumption_fbl_map = int(last_line_fbl[int(yamlConfig["fbl_ram_rw_data_start"]):int(yamlConfig["fbl_ram_rw_data_end"])].replace(" ", ""))
    else:
        grand_total_rom_consumption_fbl_map = 0
        grand_total_ram_consumption_fbl_map = 0

    consumptiondict['grand_total_rom_consumption_fbl_map'] = grand_total_rom_consumption_fbl_map
    consumptiondict['grand_total_ram_consumption_fbl_map'] = grand_total_ram_consumption_fbl_map
    return consumptiondict

# Function to process fbm.map and calculate grand rom consumption and grand ram consumption
def process_fbm_map(path_to_fbm_map):

    consumptiondict = {}

    if (path_to_fbm_map != None):
        file2 = open(path_to_fbm_map, 'r')
        Lines2 = file2.readlines()

        for line in Lines2:
            if r'Grand Total:' in line.strip():
                last_line_bootmanager = line
                break

        # Extract from grand total line ro code, rw code, ro data
        grand_total_rom_consumption_bootmanager_map = int(last_line_bootmanager[int(yamlConfig["fbm_rom_ro_code_start"]):int(yamlConfig["fbm_rom_ro_code_end"])].replace(" ", "")) + int(
            last_line_bootmanager[int(yamlConfig["fbm_rom_ro_data_start"]):int(yamlConfig["fbm_rom_ro_data_end"])].replace(" ", ""))
        grand_total_ram_consumption_bootmanager_map = int(last_line_bootmanager[int(yamlConfig["fbm_ram_rw_data_start"]):int(yamlConfig["fbm_ram_rw_data_end"])].replace(" ", ""))
    else:
        grand_total_rom_consumption_bootmanager_map = 0
        grand_total_ram_consumption_bootmanager_map = 0

    consumptiondict['grand_total_rom_consumption_fbm_map'] = grand_total_rom_consumption_bootmanager_map
    consumptiondict['grand_total_ram_consumption_fbm_map'] = grand_total_ram_consumption_bootmanager_map
    return consumptiondict

# Function to process app.map and calculate grand rom consumption and grand ram consumption
def process_app_map(path_to_app_map):

    consumptiondict = {}

    if (path_to_app_map != None):
        file2 = open(path_to_app_map, 'r')
        Lines2 = file2.readlines()

        for line in Lines2:
            if r'Grand Total:' in line.strip():
                last_line_app = line
                break

        # Extract from grand total line ro code, rw code, ro data, rw data, ro data(abs), rw data(abs)
        grand_total_rom_consumption_app_map = int(last_line_app[int(yamlConfig["app_rom_ro_code_start"]):int(yamlConfig["app_rom_ro_code_end"])].replace(" ", "")) + int(
            last_line_app[int(yamlConfig["app_rom_ro_data_start"]):int(yamlConfig["app_rom_ro_data_end"])].replace(" ", "")) + int(last_line_app[int(yamlConfig["app_rom_ro_data_abs_start"]):int(yamlConfig["app_rom_ro_data_abs_end"])].replace(" ", ""))
        grand_total_ram_consumption_app_map = int(last_line_app[int(yamlConfig["app_ram_rw_code_start"]):int(yamlConfig["app_ram_rw_code_end"])].replace(" ", "")) + int(
            last_line_app[int(yamlConfig["app_ram_rw_data_start"]):int(yamlConfig["app_ram_rw_data_end"])].replace(" ", "")) + int(last_line_app[int(yamlConfig["app_ram_rw_data_abs_start"]):int(yamlConfig["app_ram_rw_data_abs_end"])].replace(" ", ""))
    else:
        grand_total_rom_consumption_app_map = 0
        grand_total_ram_consumption_app_map = 0

    consumptiondict['grand_total_rom_consumption_app_map'] = grand_total_rom_consumption_app_map
    consumptiondict['grand_total_ram_consumption_app_map'] = grand_total_ram_consumption_app_map
    return consumptiondict


def read_from_path(path_dir, path_project, ignore_dir):

    # Open map file in read only mode
    file1 = open(path_dir, 'r')

    # append in list Lines line by line from file
    Lines = file1.readlines()

    # Find first position where to extract data
    first_position_to_extract = 0

    # Strips the newline character
    # The position can be found using link to object file generated
    for line in Lines:
        first_position_to_extract += 1
        if r'*** MODULE SUMMARY' in line.strip():
            break

    # list of file object starts after founding ***MODULE SUMMARY after 10 lines
    first_position_to_extract += 9

    # last_line is used to compute ram rom per all files ( grand total Line )
    last_line = ''

    # Find last position where data must be extracted
    last_position_to_extract = 0

    # Strips the newline character
    for line in Lines:
        last_position_to_extract += 1
        if r'Grand Total:' in line.strip():
            last_line = line
            break

    # ALL THESE LISTS ARE CREATED PER OBJECT FILE

    # extract from map file only necessary lines
    file_names = []  # Name of object files

    raw_list_ro_code = []  # Unprocessed list with ro code
    raw_list_rw_code = []  # Unprocessed list with rw_code
    raw_list_ro_data = []  # Unprocessed list with ro_data
    raw_list_rw_data = []  # Unprocessed list with rw_data
    raw_list_ro_data_abs = []  # Unprocessed list with ro_data_abs
    raw_list_rw_data_abs = []  # Unprocessed list with rw_data_abs

    # Save index position for every category so we can check their values in the map file
    for line in Lines[first_position_to_extract - 9:last_position_to_extract]:
        if "app.map" in file1.name:
            if "ro code" in line:
                ro_code = line.index("ro code")
            if "rw code" in line:
                rw_code = line.index("rw code")
            if "ro data" in line:
                ro_data = line.index("ro data")
            if "rw data" in line:
                rw_data = line.index("rw data")
            if "(abs)" in line:
                ro_data_abs = line.index("(abs)") - 2
                rw_data_abs = line.index("(abs)") + 9

        if "fbl.map" in file1.name or "fbm.map" in file1.name:
            if "ro code" in line:
                ro_code = line.index("ro code")
            if "ro data" in line:
                ro_data = line.index("ro data")
            if "rw data" in line:
                rw_data = line.index("rw data")

    # Extract only specific lines
    for line in Lines[first_position_to_extract:last_position_to_extract]:

        # Strip line first
        stripped_line = line.strip()

        # Split it after
        splitted_line = stripped_line.split()

        # Data has been filtered with IF condition
        if (len(splitted_line) > 0) and (':' not in line) and ('Linker' not in line) and ('Grand' not in line) and ('.o' in line):

            # Append the name of the file in its specific list
            file_names.append(splitted_line[0])

            # Append data in code and data lists
            if "app.map" in file1.name:
                raw_list_ro_code.append(line[ro_code:ro_code + 7])  # Extract from line information contanin only ro code
                raw_list_rw_code.append(line[rw_code:rw_code + 7])  # Extract from line information contanin only in rw code
                raw_list_ro_data.append(line[ro_data:ro_data + 7])  # Extract from line information contanin only in ro data
                raw_list_rw_data.append(line[rw_data:rw_data + 7])  # Extract from line information contanin only in rw data
                raw_list_ro_data_abs.append(line[ro_data_abs:ro_data_abs + 7])  # Extract from line information contanin only in ro data abs
                raw_list_rw_data_abs.append(line[rw_data_abs:rw_data_abs + 7])  # Extract from line information contanin only in rw data abs
            if "fbl.map" in file1.name or "fbm.map" in file1.name:
                raw_list_ro_code.append(line[ro_code:ro_code + 7])  # Extract from line information contanin only ro code
                raw_list_ro_data.append(line[ro_data:ro_data + 7])  # Extract from line information contanin only in ro data
                raw_list_rw_data.append(line[rw_data:rw_data + 7])  # Extract from line information contanin only in rw data

    # ALL THESE LISTS ARE CREATED PER OBJECT FILE
    ro_code = convert_strings_to_int(list_to_convert=raw_list_ro_code)
    rw_code = convert_strings_to_int(list_to_convert=raw_list_rw_code)
    ro_data = convert_strings_to_int(list_to_convert=raw_list_ro_data)
    rw_data = convert_strings_to_int(list_to_convert=raw_list_rw_data)
    ro_data_abs = convert_strings_to_int(list_to_convert=raw_list_ro_data_abs)
    rw_data_abs = convert_strings_to_int(list_to_convert=raw_list_rw_data_abs)

    # Compute from previous constructed lists consumption on every file
    rom_consumption_per_file = []
    if "app.map" in file1.name:
        for count_file in range(0, len(ro_code)):
            rom_consumption_per_file.append(ro_code[count_file] + ro_data[count_file] + ro_data_abs[count_file])
    else:
        for count_file in range(0, len(ro_code)):
            rom_consumption_per_file.append(ro_code[count_file] + ro_data[count_file])

    ram_consumption_per_file = []
    if "app.map" in file1.name:
        for count_file in range(0, len(rw_code)):
            ram_consumption_per_file.append(rw_code[count_file] + rw_data[count_file] + rw_data_abs[count_file])
    elif "fbl.map" in file1.name or "fbm.map" in file1.name:
        for count_file in range(0, len(rw_data)):
            ram_consumption_per_file.append(rw_data[count_file])

    # LIST USED TO EXCLUDE DIRECTORIES FROM FILE PATHS
    exclude_directories = ignore_dir

    # LIST USED TO EXCLUDE FILES FROM FILE PATHS
    exclude_files = []

    # Search all files from entire project and append them into a list
    all_files_list = []

    # THIS ROUTINE IS USED TO APPEND INTO A LIST ALL FILES FROM ENTIRE DRIVE
    for path, subdirs, files in os.walk(path_project):
        # HERE IS DEFINED HOW EXCLUDING MECHANISM WORKS
        subdirs[:] = [d for d in subdirs if d not in exclude_directories]
        files[:] = [d for d in files if d not in exclude_files]
        for name in files:
            all_files_list.append(os.path.join(path, name))

    # CONSTRUCT LIST OF FILES WITH PATH FROM ENTIRE PROJECT
    path_list = []

    # For every object file from object files list extracted from app.map
    for file in file_names:

        # First, replace .o extension with .c
        c_file = "\\" + file.replace('.o', '.c')

        # Search file in entire list of all files from project
        result = [file_name_with_path for file_name_with_path in all_files_list if c_file in file_name_with_path]

        # If found, append it into a list
        if len(result) != 0:
            path_list.append(result[0])

        # If it was not found:
        else:

            # Replace .o extension with .s extension
            s_file = "\\" + file.replace('.o', '.s')

            # Search file again in entire list of all files from project
            result = [file_name_with_path for file_name_with_path in all_files_list if s_file in file_name_with_path]

            # If found, append it into a list
            if len(result) != 0:
                path_list.append(result[0])
            # If not found, then create a section name Source file not found and append object file to it
            else:
                search_result = 'Source file not found\\' + file
                path_list.append(search_result)

    # Get all modules names from .c files
    module_names = []
    for path in path_list:
        if 'Components' in path:
            # Split all folders from path
            new_path = path.split(path_project)[1]
            new_path = new_path.split('\\')

            # Convert spitted words to list-type
            new_path = convert_elements_to_lists_elements(lists=new_path)

            # append it into a big list
            if new_path[4] not in module_names:
                module_names.append(new_path[4])


        if ('Supplier' in path) and ('src' in path) and ('T1' in path):

            # Split all folders from path
            new_path = str(path).split('\\')
            # Convert spitted words to list-type
            new_path = convert_elements_to_lists_elements(lists=new_path)

            src_index = 0
            supplier_index = 0
            for element, element_idx in zip(new_path, range(0, len(new_path))):
                if "src" in str(element):
                    src_index = element_idx
                if "Supplier" in str(element):
                    supplier_index = element_idx
            if (src_index - supplier_index ) == 2:
                # append it into a big list
                if new_path[supplier_index + 1] not in module_names:
                    module_names.append(new_path[supplier_index + 1])

        if ('plugins' in path) and ('Tools' in path) and ('src' in path) :

            # Split all folders from path
            new_path = str(path).split('\\')
            # Convert spitted words to list-type
            new_path = convert_elements_to_lists_elements(lists=new_path)

            src_index = 0
            plugin_index = 0

            for element, element_idx in zip(new_path, range(0, len(new_path))):
                if "src" in str(element):
                    src_index = element_idx
                if "plugins" in str(element):
                    plugin_index = element_idx

            if (src_index - plugin_index ) == 2:
                # append it into a big list
                if new_path[plugin_index + 1] not in module_names:
                    module_names.append(new_path[plugin_index + 1])



        if '.o' in path:
            files_not_found = ['Source file not found']
            module_names.append(files_not_found)


    for module in module_names:
        # These variables are used to compute ram and rom consumption instantaneously
        ram_consumption_per_module_value = 0
        rom_consumption_per_module_value = 0

        # Variable used to iterate in ram and rom list
        count_file = 0
        last_path = ''
        for path in path_list:
            if module[0] + '\\' in path:
                last_path = path
                # Split all folders from path
                new_path = str(path).split('\\')

                # Convert spitted words to list-type
                new_path = convert_elements_to_lists_elements(lists=new_path)

                # Include in list values of every file
                new_path[-1].append(ram_consumption_per_file[count_file])
                new_path[-1].append(rom_consumption_per_file[count_file])

                # Compute on modules values
                ram_consumption_per_module_value += ram_consumption_per_file[count_file]
                rom_consumption_per_module_value += rom_consumption_per_file[count_file]

                # append only file name in list ( LAST ELEMENT OF LIST )
                module.append(new_path[-1])

            count_file += 1

        # Append RAM AND ROM COSUMPTION PER MODULE into list AT FINAL
        module.append(ram_consumption_per_module_value)
        module.append(rom_consumption_per_module_value)

    # Eliminate dublures from list
    module_names_removed_dublures = []
    [module_names_removed_dublures.append(x) for x in module_names if x not in module_names_removed_dublures]

    # Remove source file not found element and append it at the end of the list
    index = ''
    for module_files in module_names_removed_dublures:
        for list_elements in module_files:
            if "Source file not found" == list_elements:
                index = module_names_removed_dublures.index(module_files)

    if index != '':
        module_names_removed_dublures.append(module_names_removed_dublures.pop(index))

    return module_names_removed_dublures

def replace_in_file(path_to_map_file):

    if (path_to_map_file != None):
        file = open(path_to_map_file, 'r+')
        Lines = file.readlines()
        tempLines = Lines

        line_pos = 0
        first_pos = 0
        last_pos = 0
        for line in Lines:
            line_pos += 1
            if r'*** MODULE SUMMARY' in line.strip():
                first_line = line.strip()
                first_pos = line_pos
            if r'*** ENTRY LIST' in line.strip():
                last_line = line.strip()
                last_pos = line_pos


        for line in range(first_pos, last_pos):
            if "'" in Lines[line]:
                temp = Lines[line]
                Lines[line] = temp.replace("'", " ")

        file.seek(0)
        file.truncate()
        for line in Lines:
            file.writelines(line)

def main():
    global yamlConfig

    # Read the yaml file
    with open("ram_rom_analyzer_config.yaml", 'r') as stream:
        try:
            yamlConfig= yaml.load(stream, Loader=yaml.FullLoader)
        except:
            print('No config file has been found. The program will now exit.')
            print('Perhaps YAML config file name is different. Make sure YAML config file name is: ram_rom_analyzer_config.yaml')
            return -1


    #Extract folder to ignor for Application from config file
    ignore_app_dir = yamlConfig['list_of_app_director_to_ignor']

    #Extract folder to ignor for Bootloader from config file
    ignore_fbl_dir = yamlConfig['list_of_fbl_director_to_ignor']

    #Extract folder to ignor for Bootmanager from config file
    ignore_fbm_dir = yamlConfig['list_of_fbm_director_to_ignor']

    #Extract the path for the project from config file
    path_to_project = yamlConfig['path_project']

    #Extract where the output file will be from the config file
    path_to_output = yamlConfig['output_path']

    #Set the path where the Map file for Application, Bootloader and Bootmanager are location
    path_to_app_map = str(path_to_project) + '\Release\Application\\app.map'
    path_to_fbl_map = str(path_to_project) + '\Release\Bootloader\\fbl.map'
    path_to_fbm_map = str(path_to_project) + '\Release\Bootmanager\\fbm.map'


    replace_in_file(path_to_fbl_map)
    replace_in_file(path_to_fbm_map)
    replace_in_file(path_to_app_map)


    module_names_removed_dublures_app = read_from_path(path_to_app_map, path_to_project, ignore_app_dir)
    module_names_removed_dublures_fbl = read_from_path(path_to_fbl_map, path_to_project, ignore_fbl_dir)
    module_names_removed_dublures_fbm = read_from_path(path_to_fbm_map, path_to_project, ignore_fbm_dir)

    ram_consumption_app_excel = yamlConfig['ram_consumption_app_excel']
    rom_comsumption_app_excel = yamlConfig['rom_consumption_app_excel']
    ram_consumption_fbl_excel = yamlConfig['ram_consumption_fbl_excel']
    rom_consumption_fbl_excel = yamlConfig['rom_consumption_fbl_excel']
    ram_consumption_fbm_excel = yamlConfig['ram_consumption_fbm_excel']
    rom_consumption_fbm_excel = yamlConfig['rom_consumption_fbm_excel']



    # ==================================================================================
    # WRITING IN EXCEL
    # ==================================================================================

    # Create an excel object file to write in it
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Values used to compute total rom consumption
    DFLASH_Block_Default_customer_s19 = 4096
    Unified_calibration_block = 4096
    fbl_dict = process_fbl_map(path_to_fbl_map)
    fbm_dict = process_fbm_map(path_to_fbm_map)
    app_dict = process_app_map(path_to_app_map)

    # Set Start Row Where you want to write all values
    start_row = 10

    # Set a big and nice title at the beggining
    # ==========================================================
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=7)
    sheet.cell(row=1, column=1).value = 'RAM ROM CONSUMPTION'
    sheet.cell(row=1, column=1).font = Font(bold=True, size=22)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    # ==========================================================

    # Write Grand Total Application Data Antet APP.MAP
    # ==========================================================
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
    sheet.cell(row=5, column=1).value = 'Grand Total:'
    sheet.cell(row=5, column=1).font = Font(bold=True)
    sheet.cell(row=5, column=1).alignment = Alignment(horizontal='center')

    sheet.cell(row=6, column=1).value = 'RAM Consumption (bytes):'
    sheet.cell(row=6, column=1).font = Font(bold=True)
    sheet.cell(row=6, column=1).alignment = Alignment(horizontal='center')

    sheet.cell(row=6, column=2).value = 'ROM Consumption (bytes):'
    sheet.cell(row=6, column=2).font = Font(bold=True)
    sheet.cell(row=6, column=2).alignment = Alignment(horizontal='center')

    sheet.cell(row=7, column=1).value = app_dict["grand_total_ram_consumption_app_map"]
    sheet.cell(row=7, column=1).alignment = Alignment(horizontal='center')

    sheet.cell(row=8, column=1).value = ram_consumption_app_excel
    sheet.cell(row=8, column=1).number_format = FORMAT_PERCENTAGE_00
    sheet.cell(row=8, column=1).alignment = Alignment(horizontal='center')

    sheet.cell(row=7, column=2).value = app_dict["grand_total_rom_consumption_app_map"] + Unified_calibration_block + DFLASH_Block_Default_customer_s19 + fbl_dict["grand_total_rom_consumption_fbl_map"] + fbm_dict['grand_total_rom_consumption_fbm_map']
    sheet.cell(row=7, column=2).alignment = Alignment(horizontal='center')

    sheet.cell(row=8, column=2).value = rom_comsumption_app_excel
    sheet.cell(row=8, column=2).number_format = FORMAT_PERCENTAGE_00
    sheet.cell(row=8, column=2).alignment = Alignment(horizontal='center')

    # ==========================================================

    # Write Grand Total Application Data Antet FBL.MAP
    # ==========================================================
    if (path_to_fbl_map != None):
        sheet.merge_cells(start_row=5, start_column=3, end_row=5, end_column=4)
        sheet.cell(row=5, column=3).value = 'Grand Total (fbl.map):'
        sheet.cell(row=5, column=3).font = Font(bold=True)
        sheet.cell(row=5, column=3).alignment = Alignment(horizontal='center')

        sheet.cell(row=6, column=3).value = 'RAM Consumption (bytes):'
        sheet.cell(row=6, column=3).font = Font(bold=True)
        sheet.cell(row=6, column=3).alignment = Alignment(horizontal='center')

        sheet.cell(row=6, column=4).value = 'ROM Consumption (bytes):'
        sheet.cell(row=6, column=4).font = Font(bold=True)
        sheet.cell(row=6, column=4).alignment = Alignment(horizontal='center')

        sheet.cell(row=7, column=3).value = fbl_dict["grand_total_ram_consumption_fbl_map"]
        sheet.cell(row=7, column=3).alignment = Alignment(horizontal='center')

        sheet.cell(row=8, column=3).value = ram_consumption_fbl_excel
        sheet.cell(row=8, column=3).number_format = FORMAT_PERCENTAGE_00
        sheet.cell(row=8, column=3).alignment = Alignment(horizontal='center')

        sheet.cell(row=7, column=4).value = fbl_dict["grand_total_rom_consumption_fbl_map"]
        sheet.cell(row=7, column=4).alignment = Alignment(horizontal='center')

        sheet.cell(row=8, column=4).value = rom_consumption_fbl_excel
        sheet.cell(row=8, column=4).number_format = FORMAT_PERCENTAGE_00
        sheet.cell(row=8, column=4).alignment = Alignment(horizontal='center')
    # ==========================================================

    # Write Grand Total Application Data Antet BootManager.MAP
    # ==========================================================
    if (path_to_fbm_map != None):
        sheet.merge_cells(start_row=5, start_column=5, end_row=5, end_column=6)
        sheet.cell(row=5, column=5).value = 'Grand Total (BootManager.map):'
        sheet.cell(row=5, column=5).font = Font(bold=True)
        sheet.cell(row=5, column=5).alignment = Alignment(horizontal='center')

        sheet.cell(row=6, column=5).value = 'RAM Consumption (bytes):'
        sheet.cell(row=6, column=5).font = Font(bold=True)
        sheet.cell(row=6, column=5).alignment = Alignment(horizontal='center')

        sheet.cell(row=6, column=6).value = 'ROM Consumption (bytes):'
        sheet.cell(row=6, column=6).font = Font(bold=True)
        sheet.cell(row=6, column=6).alignment = Alignment(horizontal='center')

        sheet.cell(row=7, column=5).value = fbm_dict['grand_total_ram_consumption_fbm_map']
        sheet.cell(row=7, column=5).alignment = Alignment(horizontal='center')

        sheet.cell(row=8, column=5).value = ram_consumption_fbm_excel
        sheet.cell(row=8, column=5).number_format = FORMAT_PERCENTAGE_00
        sheet.cell(row=8, column=5).alignment = Alignment(horizontal='center')

        sheet.cell(row=7, column=6).value = fbm_dict['grand_total_rom_consumption_fbm_map']
        sheet.cell(row=7, column=6).alignment = Alignment(horizontal='center')

        sheet.cell(row=8, column=6).value = rom_consumption_fbm_excel
        sheet.cell(row=8, column=6).number_format = FORMAT_PERCENTAGE_00
        sheet.cell(row=8, column=6).alignment = Alignment(horizontal='center')
    # ==========================================================

    # Write Titles in it
    # ==========================================================
    sheet.cell(row=start_row, column=1).value = 'Target'
    sheet.cell(row=start_row, column=2).value = 'Module Name'
    sheet.cell(row=start_row, column=3).value = 'RAM Consumption (module) (bytes)'
    sheet.cell(row=start_row, column=4).value = 'ROM Consumption (module) (bytes)'
    sheet.cell(row=start_row, column=5).value = 'Source file'
    sheet.cell(row=start_row, column=6).value = 'RAM Consumption (file) (bytes)'
    sheet.cell(row=start_row, column=7).value = 'ROM Consumption (file) (bytes)'

    # Set Table Header properties
    sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=1).font = Font(bold=True)
    sheet.cell(row=start_row, column=2).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=2).font = Font(bold=True)
    sheet.cell(row=start_row, column=3).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=3).font = Font(bold=True)
    sheet.cell(row=start_row, column=4).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=4).font = Font(bold=True)
    sheet.cell(row=start_row, column=5).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=5).font = Font(bold=True)
    sheet.cell(row=start_row, column=6).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=6).font = Font(bold=True)
    sheet.cell(row=start_row, column=7).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=7).font = Font(bold=True)
    sheet.cell(row=start_row + 1, column=1).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row + 1, column=1).font = Font(bold=False)

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 32
    sheet.column_dimensions['C'].width = 32
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 50
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30
    # ==========================================================

    # Define iterating variables
    sheet.cell(row=start_row + 1, column=1).value = 'Application'
    row_count = start_row + 1
    for module in module_names_removed_dublures_app:
        # Write Module name
        sheet.cell(row=row_count, column=2).value = module[0]
        sheet.cell(row=row_count, column=2).alignment = Alignment(horizontal='center')

        # Write Ram Consumption per module
        # last but one element from list is RAM Consumption per module
        sheet.cell(row=row_count, column=3).value = module[-2]
        sheet.cell(row=row_count, column=3).alignment = Alignment(horizontal='center')

        # Write Rom Consumption per module
        # Last element from list is ROM Consumption per module
        sheet.cell(row=row_count, column=4).value = module[-1]
        sheet.cell(row=row_count, column=4).alignment = Alignment(horizontal='center')

        for file in module[1:-2]:
            # Write File name
            sheet.cell(row=row_count, column=5).value = file[0]
            sheet.cell(row=row_count, column=5).alignment = Alignment(horizontal='center')

            # Write Ram Consumption per file name
            sheet.cell(row=row_count, column=6).value = file[1]
            sheet.cell(row=row_count, column=6).alignment = Alignment(horizontal='center')

            # Write Rom Consumption per file name
            sheet.cell(row=row_count, column=7).value = file[2]
            sheet.cell(row=row_count, column=7).alignment = Alignment(horizontal='center')

            row_count += 1

    row_count = row_count + 1
    sheet.cell(row=row_count, column=1).value = 'Bootloader'
    for module in module_names_removed_dublures_fbl:
        # Write Module name
        sheet.cell(row=row_count, column=2).value = module[0]
        sheet.cell(row=row_count, column=2).alignment = Alignment(horizontal='center')

        # Write Ram Consumption per module
        # last but one element from list is RAM Consumption per module
        sheet.cell(row=row_count, column=3).value = module[-2]
        sheet.cell(row=row_count, column=3).alignment = Alignment(horizontal='center')

        # Write Rom Consumption per module
        # Last element from list is ROM Consumption per module
        sheet.cell(row=row_count, column=4).value = module[-1]
        sheet.cell(row=row_count, column=4).alignment = Alignment(horizontal='center')

        for file in module[1:-2]:
            # Write File name
            sheet.cell(row=row_count, column=5).value = file[0]
            sheet.cell(row=row_count, column=5).alignment = Alignment(horizontal='center')

            # Write Ram Consumption per file name
            sheet.cell(row=row_count, column=6).value = file[1]
            sheet.cell(row=row_count, column=6).alignment = Alignment(horizontal='center')

            # Write Rom Consumption per file name
            sheet.cell(row=row_count, column=7).value = file[2]
            sheet.cell(row=row_count, column=7).alignment = Alignment(horizontal='center')

            row_count += 1

    row_count = row_count + 1
    sheet.cell(row=row_count, column=1).value = 'Bootmanager'
    for module in module_names_removed_dublures_fbm:
        # Write Module name
        sheet.cell(row=row_count, column=2).value = module[0]
        sheet.cell(row=row_count, column=2).alignment = Alignment(horizontal='center')

        # Write Ram Consumption per module
        # last but one element from list is RAM Consumption per module
        sheet.cell(row=row_count, column=3).value = module[-2]
        sheet.cell(row=row_count, column=3).alignment = Alignment(horizontal='center')

        # Write Rom Consumption per module
        # Last element from list is ROM Consumption per module
        sheet.cell(row=row_count, column=4).value = module[-1]
        sheet.cell(row=row_count, column=4).alignment = Alignment(horizontal='center')

        for file in module[1:-2]:
            # Write File name
            sheet.cell(row=row_count, column=5).value = file[0]
            sheet.cell(row=row_count, column=5).alignment = Alignment(horizontal='center')

            # Write Ram Consumption per file name
            sheet.cell(row=row_count, column=6).value = file[1]
            sheet.cell(row=row_count, column=6).alignment = Alignment(horizontal='center')

            # Write Rom Consumption per file name
            sheet.cell(row=row_count, column=7).value = file[2]
            sheet.cell(row=row_count, column=7).alignment = Alignment(horizontal='center')

            row_count += 1

    wb.save(os.path.join(path_to_output,'Ram_Rom_Analisys.xlsx'))
    print("Script ran successfully")

if __name__ == "__main__":
    main()