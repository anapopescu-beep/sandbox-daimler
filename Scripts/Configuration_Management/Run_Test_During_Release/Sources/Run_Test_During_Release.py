import subprocess
import os
import yaml
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import *
from tqdm import tqdm
import re


# ===============================================================
# FUNCTION USED TO EXTRACT A FILE NAME FROM ENTIRE PATH
# INPUT ARGUMENT: STRING
# OUTPUT ARGUMENT: STRING
# ===============================================================
def get_test_case_from_path(test_case_path):

    # Split path after \\ string
    list_of_dirs_from_path = test_case_path.split('\\')

    # Check if path has been given corectly
    if '.' in list_of_dirs_from_path[-1]:

        # return last element from splitted path
        return list_of_dirs_from_path[-1]

    # Otherwise return error code
    else:
        return 'Path not given corectly, no file found in this path!'


# ===============================================================
# FUNCTION USED TO EXTRACT ALL TESTS FROM A GIVEN PATH
# INPUT ARGUMENT: STRING, PATH TO TESTS
# OUTPUT ARGUMENT: DICTIONARY-OBJECT
# ===============================================================
def get_dictionary_with_tests_from_path(path_to_tests):

    # define dictionary with test cases
    dict_with_tests = {}

    # Exclude Directories
    exclude_directories = ['Peer_Review', 'Quality_Assurance', 'Config' , 'Test_Logs', 'Test_Cases']

    # Walk in all directories and files in given path
    for root,dirs,files in os.walk(path_to_tests):

        # Call method to exclude certan directories
        dirs[:] = [d for d in dirs if d not in exclude_directories]

        # For every founded file
        for test_case in files:

            # If a file is a test case
            if test_case.endswith('.ts'):

                # Split in folders it's root directory
                root_dir = root.split('\\')

                # For every directory in it's base path
                for dir in path_to_tests.split('\\'):

                    # Remove base directories from it's path
                    root_dir.remove(dir)

                # Get test case type from it's path
                test_case_type = root_dir[0]

                # Check if key is not already in dict
                if test_case_type not in dict_with_tests:
                    dict_with_tests[test_case_type] = []
                else:
                    # Add key in dictionary with test case
                    if test_case not in dict_with_tests[test_case_type]:
                        dict_with_tests[test_case_type].append( os.path.join(root, test_case) )
    # Return dictionary with tests
    return dict_with_tests

# ===============================================================
# FUNCTION USED TO FIND TEST CASE RESULT
# INPUT ARGUMENT: STRING, STRING
# OUTPUT ARGUMENT: STRING, PATH TO TEST RESULT
# ===============================================================
def find_test_case_result(test_case, path_to_tests):

    # Define exclude directory list
    exclude_directories = ['Peer_Review', 'Quality_Assurance', 'Config' , 'Test_Logs', 'Test_Cases', 'Test_Scripts' ]

    # Search all files from entire project and append them into a list
    all_files_list = []

    # Walk in all tests directory
    for root,dirs,files in os.walk(path_to_tests):

        # Call method to exclude certan directories
        dirs[:] = [d for d in dirs if d not in exclude_directories]

        # For every founded file in all files
        for file in files:
            all_files_list.append(os.path.join(root, file))

    # Search in all files list
    for file in all_files_list:

        # if test path is found
        if test_case.replace('.ts', '.xml').lower() in file.lower():

            # Return test path
            return file

    # Otherwise, return message
    return 'Test Result Not Found!'

# ===============================================================
# FUNCTION USED TO OPEN TEST CASE XML RESULT AND GET TEST RESULT
# INPUT ARGUMENT: STRING: PATH TO XML FILE TO PROCESS
# OUTPUT ARGUMENT: STRING: STATUS OF THE TEST
# ===============================================================
def get_result_from_xml_file(path_to_xml_file ):
    # status is read flag
    status_is_read = 0

    # initialize status as OK
    status_return = "NOK"

    # Open file for reading text.
    with open(path_to_xml_file, 'r') as myfile:

        # For each line of text
        for line in myfile:

            # verify if the line contain "Result Test:" and if the line not contain "Passed".
            if "Result Test:" in line:
                status_is_read = 1

                if 'PASSED' in line:

                    # actualize the status as "NOK"
                    status_return = "OK"

            # Exit if status has been read
            if status_is_read:
                break
    return status_return


# ===============================================================
# FUNCTION USED TO RETURN LAST CHECK IN DATE FROM PTC
# INPUT ARGUMENT: STRING, STRING
# OUTPUT ARGUMENT: STRING
# ===============================================================
def get_last_check_in_date_of_file(file_path, path_to_si ):

    # call function PTC to get member information
    batcmd = path_to_si + " memberinfo " + file_path

    # store in result output from cmd command executed
    subprocess_output = subprocess.getoutput(batcmd)

    # extract line with last check in date
    last_check_in_date_line = subprocess_output[ subprocess_output.find('Created By:') : subprocess_output.find('    State:') ].strip()

    # Remove last 2 words from string ( time zone and AM/PM time)
    try:
        last_check_in_date_line = last_check_in_date_line[ : last_check_in_date_line.rindex(' ') ]
        last_check_in_date_line = last_check_in_date_line[ : last_check_in_date_line.rindex(' ') ]
    except:
        return 'Test result file not found in PTC!'

    # Return only date from string
    last_check_in_date = last_check_in_date_line[ last_check_in_date_line.rindex('on') + 3 : last_check_in_date_line.rindex(' ') ]

    return last_check_in_date


# ===============================================================
# FUNCTION USED TO LOGIN IN PTC
# INPUT ARGUMENT: STRING, STRING
# OUTPUT ARGUMENT: NONE
# ===============================================================
def login_in_PTC(path_to_si, username, password):

    # Define flag if conected in PTC
    PTCLogIn_flag = 0

    # If not conected
    if (PTCLogIn_flag == 0):

        # Define output conected variable
        output_conected = 1

        # While not conected
        while (output_conected == 1):

            # Connect
            output_conected = os.system(path_to_si + "connect --hostname='ALVA-MKS01.alv.autoliv.int' --port='7001' --user=" + username + " --password=" + password)

        # Change flag
        PTCLogIn_flag = 1
        print('Login Succesfully in PTC account!')

# ====================================================================
# FUNCTION USED TO EXPORT DICTIONARY WITH RELEASE CONTAINTER AND DATES
# INPUT ARGUMENT: LIST, STRING
# OUTPUT ARGUMENT: DICTIONARY
# ====================================================================
def export_dictionary_from_list_of_release_containers(list_of_release_containers_id, path_to_im ):

    # Define dictionary with release container dates
    dict_with_release_containers_dates = {}

    # For ID in given list
    for release_containder_ID in list_of_release_containers_id:

        # call function PTC to get member information
        batcmd = path_to_im + " viewissue --showTimeEntries --showAnnotations  " + release_containder_ID

        # store in result output from cmd command executed
        subprocess_output = subprocess.getoutput(batcmd)

        # Extract all properties of release candidate in list
        properties_list_per_release_candidate = subprocess_output.splitlines()

        # search in all properties
        release_candidate = ''
        for property in properties_list_per_release_candidate:

            # If a release canditate found, then extract it
            if 'Release Candidate:' in property:
                release_candidate = property.replace('Release Candidate: ', '')
                break

        # Add key in dictionary
        if release_candidate not in dict_with_release_containers_dates and release_candidate != '':
            dict_with_release_containers_dates[release_candidate] = []

        # If start date found in property
        start_date = ''
        for property in properties_list_per_release_candidate:
            if 'Restraints_Release_Candidate_Start_Date:' in property:
                start_date = property.replace('Restraints_Release_Candidate_Start_Date: ', '')
                start_date = start_date.strip()
                start_date = start_date[: start_date.rindex(' ')]
                start_date = start_date[: start_date.rindex(' ')]
                break

        if start_date == '':
            print('Release Candidate: ' + release_containder_ID + ' has no start date! ')


        # If start date found in property
        build_date = ''
        for property in properties_list_per_release_candidate:
            if 'Targeted Build Date:' in property:
                build_date = property.replace('Targeted Build Date: ', '')
                build_date = build_date.strip()
                break

        if build_date == '':
            print('Release Candidate: ' + release_containder_ID + ' has no build date! ')


        # If a start date or build date has been found, then write them in dict
        if  start_date != '' and build_date != '':
            dict_with_release_containers_dates[release_candidate].append(start_date)
            dict_with_release_containers_dates[release_candidate].append(build_date)
        else:
            print('Release Candidate: ' + release_containder_ID + ' has missing dates! It will be ignored! ')
            del dict_with_release_containers_dates[release_candidate]

    # For debug purpose
    # for key, value in dict_with_release_containers_dates.items():
    #     print(key, value)

    return dict_with_release_containers_dates

# ====================================================================
# FUNCTION USED TO COMPARE A GIVEN DATE AND CHECK IF IT'S AN INTERVAL
# INPUT ARGUMENT: DICTIONARY, STRING
# OUTPUT ARGUMENT: STRING
# ====================================================================
def get_release_candidate_from_given_date(dict_with_release_containers_dates, date_to_check):

    # Define PTC Date format
    PTC_Date_Format = '%b %d, %Y'

    # Iterate in given dictionary
    for key, value in dict_with_release_containers_dates.items():

        # Convert all 3 dates to datetime-based objects
        start_date_object = datetime.strptime(value[0], PTC_Date_Format)
        build_date_object = datetime.strptime(value[1], PTC_Date_Format)
        date_to_check_object = datetime.strptime(date_to_check, PTC_Date_Format)

        # For debug purpose
        # print(start_date_object, ' - ', build_date_object, ' : ', date_to_check_object )

        if start_date_object <= date_to_check_object <= build_date_object:
            return key

    return 'Date not in release containers dates!'

# ==================================================
# FUNCTION USED TO EXTRACT FROM PTC LABEL OF FILE
# INPUT ARGUMENT: STRING, PATH
# OUTPUT ARGUMENT: STRING
# ==================================================
def get_file_label_from_PTC(file_path, path_to_si):

    # call function PTC to get member information
    batcmd = path_to_si + "memberinfo  " + file_path

    # store in result output from cmd command executed
    result = subprocess.getoutput(batcmd)

    # Extract only labels from entire string
    label_list = result[result.find('Labels:') + len('Labels:') : result.find('Change Package:')].splitlines()

    # Create list with corrected labels
    corected_label_list = []

    for label in label_list:
        # If test case not in PTC
        if 'is not a current or destined or pending member or a subsandbox' in label:
            return 'Test result file not found in PTC!'

        else:
            if label.strip() != '' and len(label.strip()) > 0:

                # Append in corected list with labels corrected label
                corected_label_list.append(label.strip())

    label_number_list = []
    for label in corected_label_list:

        # apply a regex search to extract everything between _ _ chars
        label_number_regex_result = re.search("(_(P\d\d)_)", label)

        # If found a label number, then extract it
        if label_number_regex_result != None:

            # extract label number
            label_number = label_number_regex_result.group(2)

            # Append label number in list
            if label_number not in label_number_list:
                label_number_list.append(label_number)


        else:
            # apply a regex search to extract everything between _ _ chars
            label_number_regex_result = re.search("(_(P\d\d)\W)", label)

            # If found a label number, then extract it
            if label_number_regex_result != None:

                # extract label number
                label_number = label_number_regex_result.group(2)

                # Append label number in list
                if label_number not in label_number_list:
                    label_number_list.append(label_number)

    # sort list from little to big
    label_number_list.sort(reverse=False)

    # If there are elements in sorted list
    if len(label_number_list) > 0:

        # Construct string to return
        string_to_return = 'Last run on release: ' + label_number_list[0]

    else:
        string_to_return = 'No label found for result file in PTC!'

    return string_to_return

# ===========================
# MAIN FUNCTION OF SCRIPT
# INPUT ARGUMENT: NONE
# OUTPUT ARGUMENT: NONE
# ===========================
def main():

    print("==========================================")
    print("LAST RUN OF TESTS RELEASE SCRIPT GENERATOR")
    print("Copyright Autoliv, 2021, RBE")
    print("==========================================")

    # Read yaml config file
    with open("Run_Test_During_Release_Config.yaml", 'r') as stream:
        try:
            yaml_config_file = yaml.load(stream, Loader=yaml.FullLoader)
        except:
            print('No config file has been found. The program will now exit.')
            return -1


    # LOGIN in PTC
    login_in_PTC(path_to_si=yaml_config_file['Path to si.exe'] , username= yaml_config_file['PTC Username'], password=yaml_config_file['PTC Password'] )

    # Extract dictionary with values from tests
    dict_with_tests = get_dictionary_with_tests_from_path( path_to_tests=yaml_config_file['Path to tests'] )

    # Ignore some test types
    for test_type_to_ignore in yaml_config_file['List of test types to ignore']:
        del dict_with_tests[test_type_to_ignore]

    # for key, value in dict_with_tests.items():
    #     print(key, ' : ', value)

    # Extract dictionary with release container dates
    dict_with_release_containers_dates = export_dictionary_from_list_of_release_containers(list_of_release_containers_id= yaml_config_file['List of release containers'], path_to_im=yaml_config_file['Path to im.exe'])

    # Create an excel object file to write in it
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Set Start Row Where you want to write all values
    start_row = 5

    # Set a big and nice title at the beggining
    # ==========================================================
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=6)
    sheet.cell(row=1, column=1).value = 'TESTS LIST AND LAST RUN RELEASE NUMBER'
    sheet.cell(row=1, column=1).font = Font(bold=True,  size=22)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    # ==========================================================

    # Write Titles in it
    # ==========================================================
    sheet.cell(row=start_row, column=1).value = 'Test Case Name'
    sheet.cell(row=start_row, column=2).value = 'Test Case Result'
    sheet.cell(row=start_row, column=3).value = 'Last CHECK IN Date'
    sheet.cell(row=start_row, column=4).value = 'Release found by Release Container list'
    sheet.cell(row=start_row, column=5).value = 'Release found by Label in PTC'

    # Set Table Header properties
    sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=1).font = Font(bold=True)
    sheet.cell(row=start_row, column=2).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=2).font = Font(bold=True)
    sheet.cell(row=start_row, column=3).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=3).font = Font(bold=True)
    sheet.cell(row=start_row, column=4).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=4).font = Font(bold=True)
    sheet.cell(row=start_row, column=5).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=5).font = Font(bold=True)

    sheet.column_dimensions['A'].width = 70
    sheet.column_dimensions['B'].width = 32
    sheet.column_dimensions['C'].width = 32
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 50

    # ==========================================================

    # Get to next row
    start_row += 1

    for key, value in dict_with_tests.items():

        print('\n', '------- Analyse:', key, ' -------', '\n')

        # Write Test type in document
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        sheet.cell(row=start_row, column=1).value = key
        sheet.cell(row=start_row, column=1).font = Font(bold=True)
        sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')

        # Get to next row
        start_row += 1

        for test_case_with_path, list_increment in zip( value, tqdm(value) ):

            # Get test case name
            test_case_name = get_test_case_from_path(test_case_with_path)

            # Write Test case name
            sheet.cell(row=start_row, column=1).value = test_case_name.replace('.ts', '')
            sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')

            # Get path of test result
            test_result_with_path = find_test_case_result( test_case=test_case_name, path_to_tests=yaml_config_file['Path to tests'] )

            if (test_result_with_path != 'Test Result Not Found!' ):

                # Get test result label from PTC
                last_run_label = get_file_label_from_PTC(file_path = test_result_with_path , path_to_si = yaml_config_file['Path to si.exe'] )

                # get result from path with test result
                result = get_result_from_xml_file(test_result_with_path)

                # Write test case result value
                sheet.cell(row=start_row, column=2).value = result
                sheet.cell(row=start_row, column=2).alignment = Alignment(horizontal='center')

                # Get date of last check in of result
                last_check_in = get_last_check_in_date_of_file( file_path=test_result_with_path, path_to_si=yaml_config_file['Path to si.exe'] )

                # Write test case result last date
                sheet.cell(row=start_row, column=3).value = last_check_in
                sheet.cell(row=start_row, column=3).alignment = Alignment(horizontal='center')

                if last_check_in != 'Test result file not found in PTC!':
                    # Get release containter from given date
                    release_container = get_release_candidate_from_given_date(dict_with_release_containers_dates=dict_with_release_containers_dates, date_to_check=last_check_in)
                else:
                    release_container = '-'

                # Write test case result last date
                sheet.cell(row=start_row, column=4).value = release_container
                sheet.cell(row=start_row, column=4).alignment = Alignment(horizontal='center')

                # Write test case result last date
                sheet.cell(row=start_row, column=5).value = last_run_label
                sheet.cell(row=start_row, column=5).alignment = Alignment(horizontal='center')

            # If test result has not been found
            else:

                # Write test case result value
                sheet.cell(row=start_row, column=2).value = test_result_with_path
                sheet.cell(row=start_row, column=2).alignment = Alignment(horizontal='center')

                sheet.cell(row=start_row, column=3).value = '-'
                sheet.cell(row=start_row, column=3).alignment = Alignment(horizontal='center')

                sheet.cell(row=start_row, column=4).value = '-'
                sheet.cell(row=start_row, column=4).alignment = Alignment(horizontal='center')

            start_row += 1

    wb.save(yaml_config_file['File name and path'])

if __name__ == '__main__':
    main()