"""peer_review_and_warningReport_analysis.py: Used to generate an xlsx with statistics data
Parameters has been given from argparse command line interface.
As input, it is needed an path where components folder is ( S/Components/Application ),
and optionally, strings with build_number and date
As output, it is needed a path to generate created excel report
"""

__author__ = 'Ardeleanu Lucian'
__copyright__ = "Copyright 2021, Autoliv Electronic"
__version__ = "1.0"
__email__ = 'lucian.ardeleanu@autoliv.com'
__status__ = 'Released'

# ============ IMPORTS ====================
import argparse
import os
import glob
import subprocess
import string
from natsort import natsorted

import xlrd

import openpyxl
from openpyxl.styles import *
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import re

import csv
from PIL import Image
import matplotlib.pyplot as plt
import numpy as np
import yaml

# ===============================================================
# FUNCTION TO CONVERT COLUMN NUMBER INTO LETTERS
# INPUT: INTEGER NUMBER
# OUTPUT: STRING: LETTERS ( FOR EXCEL COLUMNS )
# ===============================================================
def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

# ===============================================================
# FUNCTION USED TO OPEN EXCEL FILE AND EXTRACT DATA FROM IT
# INPUT ARGUMENT: STRING: PATH TO EXCEL FILE TO PROCESS
# OUTPUT ARGUMENT: LIST OF STRINGS WITH DATA FROM EXCEL
# ===============================================================
def extract_data_from_excel(file_to_process):

    # Define List in witch data will be stored
    raw_list = []


    # If file ends with xls, open it with xrdp
    if file_to_process.endswith(".xls"):
        if 'Peer Review' in file_to_process or 'Peer' in file_to_process or 'Review' in file_to_process:

            # open workbook
            wb = xlrd.open_workbook(file_to_process)
            sheets = wb.sheet_names()

            # open sheet
            # check which file is needed for opening

            for sheet_name in sheets:
                if ( 'code review' in sheet_name.lower() or 'review' in sheet_name.lower() or 'code' in sheet_name.lower() or sheet_name.lower() == 'review' ):
                    ws = wb.sheet_by_name(sheet_name)
                    break

        else:
            wb = xlrd.open_workbook(file_to_process)
            ws = wb.sheet_by_index(0)

        # Get total number of rows and columns
        rows = ws.nrows

        # We don't actually need all columns for document
        #columns = ws.ncols
        columns = 4

        # append document data in list
        for i in range(0, rows):
            data_from_row = []
            for j in range(0, columns):
                data_from_row.append(ws.cell_value(i,j))
            raw_list.append(data_from_row)

    # if file ends with xlsx, open it with openpyxl
    else:
        if 'Peer Review' in file_to_process or 'Peer' in file_to_process or 'Review' in file_to_process:

            # Open workbook
            wb = openpyxl.load_workbook(file_to_process)
            sheets = wb.sheetnames
            # open sheet

            for sheet_name in sheets:
                if ( 'code review' in sheet_name.lower() or 'review' in sheet_name.lower() or 'code' in sheet_name.lower() or sheet_name.lower() == 'review' ):
                    ws = wb[sheet_name]

                    break
        else:
            wb = openpyxl.load_workbook(file_to_process)
            ws = wb.active

        # Get all rows
        rows = ws.max_row

        # if needed, get all columns
        #columns = max_column
        columns = 5

        # append document data in list
        for i in range(1, rows):
            data_from_row = []
            for j in range(1, columns):
                data_from_row.append(ws.cell(row=i, column=j).value)
            raw_list.append(data_from_row)

    # Convert raw list to list with strings
    data_from_workbook = []
    for row in raw_list:
        rows = []
        for element in row:
            rows.append(str(element))
        data_from_workbook.append(rows)

    return data_from_workbook

# ================================================================================================
# FUNCTION USED TO FIND A WARNING REPORT DOCUMENT AND EXTRACT SOURCE FILE MEMBER REVISION FROM IT
# INPUT ARGUMENT: STRING: PATH TO COMPONENTS, STRING:MODULE NAME, STRING: SOURCE FILE
# OUTPUT ARGUMENT: STRING: MEMBER REVISION
# =================================================================================================
def find_Warning_Report_document_and_extract_member_revision(path_to_modules, module, source_file):
    # SEARCH WARNING REPORT WORKBOOK IN ENTIRE FOLDER MODULE
    # Set a precondition of found Peer Review Workbook file

    found = False
    file_to_process = ''
    for root, dirs, files in os.walk(path_to_modules + '/' + module + '\Quality_Assurance\Static_Analysis' ):
        # exclude directories if needed and files
        for file in files:

            if (file.endswith(".xls") or file.endswith(".xlsm") or file.endswith(".xlsx") ) and ('WarningReport' in file) and ( source_file in file):
                file_to_process = os.path.join(root, file)
                found = True
                break
    if source_file.endswith(".c"):
        if found == True:
            # Oepn Workbook
            data_from_workbook = extract_data_from_excel(file_to_process)

            # Define initial member revision state
            member_version = ''
            # Iterating throu elements
            for row in data_from_workbook:
                for element in row:
                    if ('Warning' in element) and (source_file in element):

                        # find where source file starts in string
                        start_of_source_file_name = element.index(source_file)

                        # extract member revision imediately after source file name and end of string
                        member_version = element[start_of_source_file_name + len(source_file) : ].replace(" ", "")
            return member_version

        # if not found any relevant document, return message
        else:
            return 'WARNING REPORT DOCUMENT NOT FOUND'
    else:
        return 'NOT APPLICABLE'

# ================================================================================================
# FUNCTION USED TO FIND A PEER REVIEW WORKBOOK DOCUMENT AND EXTRACT SOURCE FILE MEMBER REVISION FROM IT
# ( ONLY REVIEW SHEET )
# INPUT ARGUMENT: STRING: PATH TO COMPONENTS, STRING:MODULE NAME, STRING: SOURCE FILE
# OUTPUT ARGUMENT: STRING: MEMBER REVISION
# =================================================================================================
def get_member_revision_number_from_peer_review_workbook_review_sheet(path_to_modules,module, source_file):
    # SEARCH PEER REVIEW WORKBOOK IN ENTIRE FOLDER MODULE
    # Set a precondition of found Peer Review Workbook file
    found = False

    file_to_process = ''
    for root, dirs, files in os.walk(path_to_modules + '/' + module + '\Quality_Assurance\Peer_Review' ):
        # exclude directories if needed and files
        for file in files:

            if (file.endswith(".xls") or file.endswith(".xlsm") or file.endswith(".xlsx") ) and ('Peer Review' in file):
                file_to_process = os.path.join(root, file)
                found = True
                break
    if found == True:
        # extract data from excel
        data_from_workbook = extract_data_from_excel(file_to_process)

        # List needed to select only useable source file versions
        source_file_version_list = []


        # Exctact source file version
        for row in data_from_workbook:
            for element in row[0:3]:

                # if source file is found in list
                if element.lower() == source_file.lower() or source_file + ' ' in element:

                    # Get source file version
                    source_file_version = row[row.index(element) - 1]

                    # Replace blank spaces from source file version
                    source_file_version = source_file_version.replace(" ", "")
                    source_file_version = source_file_version.replace("..", ".")

                    # apply some filtering to data and check if version not already appended to versions list
                    if '.' in source_file_version and source_file_version not in source_file_version_list:
                        source_file_version_list.append(source_file_version)


        # Sort Naturaly file version list
        if yaml_config_file["sort_version"]=="natsorted":
            sorted_source_file_version_list = natsorted(source_file_version_list)
        else:
            sorted_source_file_version_list = source_file_version_list



        if len(sorted_source_file_version_list) != 0:
            return sorted_source_file_version_list[-1]
        else:
            return 'FILE NOT FOUND IN PEER REVIEW'
    else:
        return 'PEER REVIEW NOT FOUND FOR THIS MODULE'

# ================================================================================================
# FUNCTION USED TO FIND A PEER REVIEW WORKBOOK DOCUMENT AND EXTRACT SOURCE FILE MEMBER REVISION FROM IT
# ( ONLY CODE CHECKLIST SHEET )
# INPUT ARGUMENT: STRING: PATH TO COMPONENTS, STRING:MODULE NAME, STRING: SOURCE FILE
# OUTPUT ARGUMENT: STRING: MEMBER REVISION
# =================================================================================================
def get_member_revision_number_from_peer_review_workbook_checklist_sheet(path_to_modules,module, source_file, lower_case_sheet):
    # SEARCH PEER REVIEW WORKBOOK IN ENTIRE FOLDER MODULE
    # Set a precondition of found Peer Review Workbook file
    found = False
    file_to_process = ''
    for root, dirs, files in os.walk(path_to_modules + '/' + module + '\Quality_Assurance\Peer_Review' ):
        # exclude directories if needed and files
        for file in files:

            if (file.endswith(".xls") or file.endswith(".xlsm") or file.endswith(".xlsx") ) and ('Peer Review' in file):
                file_to_process = os.path.join(root, file)
                found = True
                break

    if found == True:

        # Define List in witch data will be stored
        raw_list = []

        # If file ends with xls, open it with xrdp
        if file_to_process.endswith(".xls"):

            # open workbook
            wb = xlrd.open_workbook(file_to_process)
            sheets = wb.sheet_names()

            # Variable to check if code checklist sheet is found
            code_checklist_found_flag = 0
            for sheet_name in sheets:

                if ( lower_case_sheet in sheet_name.lower() or lower_case_sheet == sheet_name.lower() ):
                    ws = wb.sheet_by_name(sheet_name)
                    code_checklist_found_flag = 1
                    break
                else:
                    code_checklist_found_flag = 0

            if code_checklist_found_flag == 0:
                return 'CHECKLIST SHEET NOT FOUND'

            # Configure rows and columns to read exactly what is needed
            rows = 4
            columns = ws.ncols

            # append document data in list
            for i in range(0, rows):
                data_from_row = []
                for j in range(0, columns):
                    data_from_row.append(ws.cell_value(i, j))
                raw_list.append(data_from_row)


        # if file ends with xlsx, open it with openpyxl
        else:
            # Open workbook
            wb = openpyxl.load_workbook(file_to_process)
            sheets = wb.sheetnames

            # open sheet
            # Variable to check if code checklist sheet is found
            code_checklist_found_flag = 0
            for sheet_name in sheets:

                if ( lower_case_sheet in sheet_name.lower() or lower_case_sheet == sheet_name.lower() ):
                    ws = wb[sheet_name]
                    code_checklist_found_flag = 1
                    break
                else:
                    code_checklist_found_flag = 0
            if code_checklist_found_flag == 0:
                return 'CHECKLIST SHEET NOT FOUND'

            # Configure rows and columns to read exactly what is needed
            rows = 5
            columns = ws.max_column

            # append document data in list
            for i in range(1, rows):
                data_from_row = []
                for j in range(1, columns):
                    data_from_row.append(ws.cell(row=i, column=j).value)
                raw_list.append(data_from_row)


        # Convert raw list to list with strings
        data_from_workbook = []
        for row in raw_list:
            rows = []
            for element in row:
                rows.append(str(element))
            data_from_workbook.append(rows)

        # List needed to select only useable source file versions
        source_file_version_list = []

        # Exctact source file version
        for row in data_from_workbook:
            for element in row:
                # if source file is found in list
                if source_file in element:
                    # Search version in entire string
                    try:
                        source_file_version = re.search('(\d.+)', element.split("-")[1]).group(1)
                    except:
                        print('Source file version not found for', element, ' in code checklist sheet!')
                        source_file_version = "NONE"

                    # apply some filtering to data and check if version not already appended to versions list
                    source_file_version = source_file_version.replace(" ", "")
                    source_file_version = source_file_version.replace("..", ".")
                    source_file_version = source_file_version.replace(",", "")

                    if '.' in source_file_version and source_file_version not in source_file_version_list:
                        source_file_version_list.append(source_file_version)

        # Sort Naturaly file version list
        if yaml_config_file["sort_version"] == "natsorted":
            sorted_source_file_version_list = natsorted(source_file_version_list)
        else:
            sorted_source_file_version_list = source_file_version_list

        if len(sorted_source_file_version_list) != 0:
            return sorted_source_file_version_list[-1]
        else:
            return 'FILE NOT FOUND IN CHECK LIST SHEET'
    else:
        return 'PEER REVIEW NOT FOUND FOR THIS MODULE'

# ================================================================================================
# FUNCTION USED TO GET MEMBER REVISION NUMBER FROM PTC
# INPUT ARGUMENT: STRING: PATH TO FILE
# OUTPUT ARGUMENT: STRING: MEMBER REVISION
# =================================================================================================
def get_member_revision_number_from_PTC(path):

    # Global path to si.exe ( if is not exist in bench )
    path_to_si = '"C:/Program Files (x86)/Integrity/ILMClient11/bin/si.exe" '

    # call function PTC to get member information
    batcmd = path_to_si + " memberinfo " + '"' + path + '"'

    # store in result output from cmd command executed
    result = subprocess.getoutput(batcmd)

    # search for member revision substring in all string
    substring_start = result.find('Member Revision:')
    substring_end = result.find('Created')

    # retain in variable member revision only member revision finded

    member_revision = result[substring_start + 17:substring_end]

    if  'is not a current or destined or pending member or a subsandbox' in member_revision:
        return 'SOURCE FILE NOT FOUND IN PTC'

    else:
        member_revision = member_revision.replace(r' ','')
        member_revision = member_revision.rstrip()
        return member_revision

# ================================================================================================
# FUNCTION USED TO PLOT STATISTICS PER EVERY MODULE IN PNG FILE
# INPUT ARGUMENT: LISTS WITH STRINGS
# OUTPUT ARGUMENT: FILE: GENERATED PNG FILE WITH NAME FROM GIVEN ARGUMENT ( OUTPUT PATH )
# =================================================================================================
def plot_statistics_per_module(modules, list_of_number_of_not_ok_source_files_per_module, output_path):
    # Plot is created using data processed up
    bars = list_of_number_of_not_ok_source_files_per_module
    height = list_of_number_of_not_ok_source_files_per_module

    # Set plot title and path
    plot_title = 'Statistics of NOK Source file versions per module'
    path_to_save_graph = os.path.join(output_path, 'Peer_Review_and_warningReport_analysis_per_module.PNG')

    # Process data
    y_pos = np.arange(len(bars))
    plt.figure(figsize=(20, 10))
    plt.bar(y_pos, height)

    # Set names on X side
    names = modules
    plt.xticks(y_pos, names, rotation=90)

    # Thus we have to give more margin:
    plt.subplots_adjust(bottom=0.4)

    # It's the same concept if you need more space for your titles
    plt.title(plot_title)
    plt.subplots_adjust(top=0.7)
    plt.savefig(path_to_save_graph)

    # Process image after it is created to scale it to certain dimension
    im = Image.open(path_to_save_graph)
    crop_rectangle = (200, 270, 1810, 800)  # left, top, right, bottom
    cropped_im = im.crop(crop_rectangle)
    newsize = (1450, 400)
    imresize = cropped_im.resize(newsize)

    im = imresize.save(path_to_save_graph)

# ================================================================================================
# FUNCTION USED TO PLOT STATISTICS PER EVERY BUILD IN PNG FILE
# INPUT ARGUMENT: LISTS WITH STRINGS
# OUTPUT ARGUMENT: FILE: GENERATED PNG FILE WITH NAME FROM GIVEN ARGUMENT ( OUTPUT PATH )
# =================================================================================================
def plot_statistics_per_run(csv_list, output_path):

    # Define lists with data
    build_number_and_date = []
    number_of_nok_source_files_list = []

    # for last 100 elements in list
    for row in csv_list[-100:]:
        number_of_nok_source_files, build_number, date = row[0:3]
        number_of_nok_source_files_list.append(int(number_of_nok_source_files))
        build_number_and_date.append(build_number + ' - ' + date)


    # Plot is created using data processed up
    bars = number_of_nok_source_files_list
    height = number_of_nok_source_files_list

    # Set plot title and path
    plot_title = 'Statistics of all NOK Source file versions per date and build number'
    path_to_save_graph = os.path.join(output_path, 'Peer_Review_and_warningReport_analysis_per_date_and_build.PNG')

    # Process data
    y_pos = np.arange(len(bars))
    plt.figure(figsize=(20, 10))
    plt.bar(y_pos, height)

    # Set names on X side
    names = build_number_and_date
    plt.xticks(y_pos, names, rotation=90)

    # Thus we have to give more margin:
    plt.subplots_adjust(bottom=0.4)

    # It's the same concept if you need more space for your titles
    plt.title(plot_title)
    plt.subplots_adjust(top=0.7)
    plt.savefig(path_to_save_graph)

    # Process image after it is created to scale it to certain dimension
    im = Image.open(path_to_save_graph)
    crop_rectangle = (200, 270, 1810, 800)  # left, top, right, bottom
    cropped_im = im.crop(crop_rectangle)
    newsize = (1450, 400)
    imresize = cropped_im.resize(newsize)

    im = imresize.save(path_to_save_graph)

# ===============================================================
# MAIN FUNCTION OF SCRIPT
# INPUT ARGUMENT: NONE
# OUTPUT ARGUMENT: NONE
# ===============================================================
def main():
    global yaml_config_file

    parser = argparse.ArgumentParser()
    # parser.add_argument('-i', '--input-path', help="Path to the components directory, E.G. S:\Components\Application but from C: drive", required=True)
    parser.add_argument('-i', '--input-yaml-config-path', help="Path to source_file_version_list", required=True)
    #parser.add_argument('-o', '--output-path', help="Path to the output files, E.G. C:\Output_folder", required=False)
    parser.add_argument('-d', '--data', help="Data to fill in global_warning_stats.csv", required=False)
    parser.add_argument('-b', '--build-number', help="Build number to fill in global_warning_stats.csv", required=False)
    #parser.add_argument('-p', '--ptc-sortversion', help="How to sort yaml", required=False)
    args = parser.parse_args()

    yaml_config_path = os.path.join(args.input_yaml_config_path, "PRWR_analisys_configuration_file.yaml")

    print(yaml_config_path)

    print('Peer Review and warningReport Analysis Script')

    # Read yaml config file
    with open(yaml_config_path, 'r') as stream:
        try:
            yaml_config_file = yaml.load(stream, Loader=yaml.FullLoader)
        except Exception as e:
            print('No config file has been found. The program will now exit.')
            print('Error encoutered in reading yaml config file: ', e)
            return -1

    # Extract file names to ignore from config file
    files_to_ignore_list = yaml_config_file['list_of_files_to_ignore']
    # Extract folder names to ignore from config file
    folders_to_ignore_list = yaml_config_file['list_of_directories_to_ignore']
    path_to_modules = yaml_config_file['sw_comp_folder']
    output_path = yaml_config_file['output_folder']
    sort_version = yaml_config_file['sort_version']
    sw_components_list = yaml_config_file['sw_components_list']


    if files_to_ignore_list is None:
        files_to_ignore_list = [""]

    if folders_to_ignore_list is None:
        folders_to_ignore_list =["random_string_not_in_path"]

    exclude_directories = folders_to_ignore_list
    exclude_files = files_to_ignore_list

    # Get a list of all modules in a directory
    modules_list = os.listdir(path_to_modules)

    # Create an excel object file to write in it
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Set Start Row Where you want to write all values
    start_row = 4

    # Set a big and nice title at the beggining
    # ==========================================================
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=6)
    sheet.cell(row=1, column=1).value = 'PEER REVIEW AND WARNING REPORT ANALISYS'
    sheet.cell(row=1, column=1).font = Font(bold=True,  size=18)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    # ==========================================================

    # Write Titles in it
    # ==========================================================
    sheet.cell(row=start_row, column=1).value = 'Source File'
    sheet.cell(row=start_row, column=2).value = 'PTC Version'
    sheet.cell(row=start_row, column=3).value = 'Peer Review Workbook Version'
    sheet.cell(row=start_row, column=4).value = 'QAC MISRA Report Version'
    sheet.cell(row=start_row, column=5).value = 'Code Checklist Version'
    sheet.cell(row=start_row, column=6).value = 'UTSR Checklist Version'
    sheet.cell(row=start_row, column=7).value = 'SDD Checklist Version'
    sheet.cell(row=start_row, column=8).value = 'ITSR Checklist Version'


    # Set Table Header properties
    sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=1).font = Font(bold=True)
    sheet.cell(row=start_row, column=2).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=2).font = Font(bold=True)
    sheet.cell(row=start_row, column=3).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=3).font = Font(bold=True)
    sheet.cell(row=start_row, column=4).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=4).font = Font(bold=True)
    sheet.cell(row=start_row, column=5).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=5).font = Font(bold=True)
    sheet.cell(row=start_row, column=6).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=6).font = Font(bold=True)
    sheet.cell(row=start_row, column=7).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=7).font = Font(bold=True)
    sheet.cell(row=start_row, column=8).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=8).font = Font(bold=True)

    # freeze first rows
    sheet.freeze_panes = 'A' + str(start_row + 1 )

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 32
    sheet.column_dimensions['C'].width = 42
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 42
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30
    # ==========================================================

    # Define from where to start
    row_count = start_row + 1

    # Define list with not ok source files per module
    list_of_number_of_not_ok_source_files_per_module = []

    # For every module in list
    modules_list_for_plot = []
    for module in modules_list[0:]:
        if (module in sw_components_list or "all" in sw_components_list) and (module not in exclude_directories):
            modules_list_for_plot.append(module)

            os.system('echo Analyse module: ' + module)
            # Define a counter needed for ploting graphs
            number_of_not_ok_source_files_module = 0

            number_of_not_ok_source_files_ptc_per_module = 0
            number_of_not_ok_source_files_peer_review_per_module = 0
            number_of_not_ok_source_files_warning_report_per_module = 0
            number_of_not_ok_source_files_code_check_list_per_module = 0
            number_of_not_ok_source_files_sdd_check_list_per_module = 0
            number_of_not_ok_source_files_itsr_check_list_per_module = 0
            number_of_not_ok_source_files_utsr_check_list_per_module = 0

            # Write module name and apply some stiling to it
            sheet.merge_cells(start_row=row_count, start_column=1, end_row=row_count, end_column=4)
            sheet.cell(row=row_count, column=1).alignment = Alignment(horizontal='center')
            sheet.cell(row=row_count, column=1).font = Font(bold=True)
            sheet.cell(row=row_count, column=1).value = module

            # Increment to next row
            row_count += 1

            for root, dirs, files in os.walk(path_to_modules + '/' + module ):
                # exclude directories if needed and files
                # dirs[:] = [d for d in dirs if d not in exclude_directories]
                files[:] = [d for d in files if d not in exclude_files]

                ignore_folder_flag = False
                for ignored_folder in folders_to_ignore_list:
                    if ignored_folder in root.replace(path_to_modules, ''):
                        ignore_folder_flag = True
                        break
                if not ignore_folder_flag:
                    for file in files:
                        # If found a source file
                        if (file.endswith(".c") or file.endswith(".h") or file.endswith(".tmb") or '_SW_DetailedDesign' in file or 'Integration Test Specification' in file):

                            # Save file path and file name in variables
                            path_to_source_file = os.path.join(root, file)
                            source_file_name = file

                            member_revision_from_ptc = get_member_revision_number_from_PTC(path_to_source_file)

                            member_revision_peer_review = get_member_revision_number_from_peer_review_workbook_review_sheet(path_to_modules=path_to_modules, module=module, source_file= source_file_name)
                            member_revision_warning_report = find_Warning_Report_document_and_extract_member_revision(path_to_modules=path_to_modules, module= module, source_file=source_file_name)

                            if ( ( file.endswith(".c") or file.endswith(".h") ) and ('Implementation' in path_to_source_file) ):
                                member_revision_code_check_list = get_member_revision_number_from_peer_review_workbook_checklist_sheet(path_to_modules=path_to_modules, module= module, source_file=source_file_name, lower_case_sheet='code checklist')
                            else:
                                member_revision_code_check_list = 'NOT APPLICABLE'

                            if file.endswith(".tmb"):
                                member_revision_utsr_list = get_member_revision_number_from_peer_review_workbook_checklist_sheet(path_to_modules=path_to_modules, module= module, source_file=source_file_name, lower_case_sheet='utsr checklist')
                            else:
                                member_revision_utsr_list = 'NOT APPLICABLE'

                            if '_SW_DetailedDesign' in file:
                                member_revision_sdd_check_list = get_member_revision_number_from_peer_review_workbook_checklist_sheet(path_to_modules=path_to_modules, module= module, source_file=source_file_name, lower_case_sheet='sdd checklist')
                            else:
                                member_revision_sdd_check_list = 'NOT APPLICABLE'

                            if 'Integration Test Specification' in file:
                                member_revision_itsr_check_list = get_member_revision_number_from_peer_review_workbook_checklist_sheet(path_to_modules=path_to_modules, module= module, source_file=source_file_name, lower_case_sheet='itsr checklist')
                            else:
                                member_revision_itsr_check_list = 'NOT APPLICABLE'

                            #print('Source File: ' + source_file_name + ': PTC: ' +  member_revision_from_ptc + ' PEER REVIEW: ' + member_revision_peer_review + ' WARNING REPORT: ' + member_revision_warning_report)

                            # write in report versions
                            sheet.cell(row=row_count, column=1).value = source_file_name
                            sheet.cell(row=row_count, column=2).value = member_revision_from_ptc
                            sheet.cell(row=row_count, column=3).value = member_revision_peer_review
                            sheet.cell(row=row_count, column=4).value = member_revision_warning_report
                            sheet.cell(row=row_count, column=5).value = member_revision_code_check_list
                            sheet.cell(row=row_count, column=6).value = member_revision_utsr_list
                            sheet.cell(row=row_count, column=7).value = member_revision_sdd_check_list
                            sheet.cell(row=row_count, column=8).value = member_revision_itsr_check_list

                            if (member_revision_from_ptc == member_revision_peer_review == member_revision_warning_report == member_revision_code_check_list ) or \
                                (member_revision_from_ptc == member_revision_peer_review == member_revision_warning_report == member_revision_utsr_list) or \
                                (member_revision_from_ptc == member_revision_peer_review == member_revision_warning_report == member_revision_sdd_check_list) or \
                                (member_revision_from_ptc == member_revision_peer_review == member_revision_warning_report == member_revision_itsr_check_list) or \
                                (member_revision_from_ptc == member_revision_peer_review == member_revision_code_check_list and  member_revision_warning_report == 'NOT APPLICABLE' ) or \
                                (member_revision_from_ptc == member_revision_peer_review == member_revision_utsr_list and member_revision_warning_report == 'NOT APPLICABLE') or \
                                (member_revision_from_ptc == member_revision_peer_review == member_revision_sdd_check_list and member_revision_warning_report == 'NOT APPLICABLE') or \
                                (member_revision_from_ptc == member_revision_peer_review == member_revision_itsr_check_list and member_revision_warning_report == 'NOT APPLICABLE'):

                                # Color row if found same version
                                for rows in sheet.iter_rows(min_row=row_count, max_row=row_count, min_col=1, max_col=8):
                                    for cell in rows:
                                        cell.fill = PatternFill(start_color='06f800', end_color='06f800', fill_type="solid")

                            if( member_revision_from_ptc == 'SOURCE FILE NOT FOUND IN PTC' ):
                                number_of_not_ok_source_files_ptc_per_module += 1

                            if member_revision_peer_review !=  member_revision_from_ptc:
                                number_of_not_ok_source_files_peer_review_per_module += 1

                            if (member_revision_code_check_list !=  member_revision_from_ptc) and (member_revision_code_check_list != 'NOT APPLICABLE'):
                                number_of_not_ok_source_files_code_check_list_per_module += 1

                            if (member_revision_utsr_list !=  member_revision_from_ptc) and (member_revision_utsr_list != 'NOT APPLICABLE'):
                                number_of_not_ok_source_files_utsr_check_list_per_module += 1

                            if (member_revision_sdd_check_list !=  member_revision_from_ptc) and (member_revision_sdd_check_list != 'NOT APPLICABLE'):
                                number_of_not_ok_source_files_sdd_check_list_per_module += 1

                            if (member_revision_itsr_check_list !=  member_revision_from_ptc) and (member_revision_itsr_check_list != 'NOT APPLICABLE'):
                                number_of_not_ok_source_files_itsr_check_list_per_module += 1

                            if ( member_revision_warning_report !=  member_revision_from_ptc ) and member_revision_warning_report != 'NOT APPLICABLE' :
                                number_of_not_ok_source_files_warning_report_per_module += 1

                            number_of_not_ok_source_files_module = number_of_not_ok_source_files_code_check_list_per_module + \
                                                                   number_of_not_ok_source_files_utsr_check_list_per_module + \
                                                                   number_of_not_ok_source_files_sdd_check_list_per_module + \
                                                                   number_of_not_ok_source_files_itsr_check_list_per_module + \
                                                                   number_of_not_ok_source_files_warning_report_per_module + \
                                                                   number_of_not_ok_source_files_peer_review_per_module + \
                                                                   number_of_not_ok_source_files_ptc_per_module


                            # increment row
                            row_count += 1

            list_of_number_of_not_ok_source_files_per_module.append(number_of_not_ok_source_files_module)
    sum_of_all_NOK_source_files_per_all_modules = sum(list_of_number_of_not_ok_source_files_per_module)

    if args.build_number != None and args.data != None :
        if not os.path.isfile(os.path.normpath(os.path.join(output_path, 'Peer_Review_and_warningReport_analysis_history.csv'))):
            global_file_obj = open(os.path.normpath(os.path.join(output_path, 'Peer_Review_and_warningReport_analysis_history.csv')), 'w+',  newline='')
        else:
            global_file_obj = open(os.path.normpath(os.path.join(output_path, 'Peer_Review_and_warningReport_analysis_history.csv')), 'a+',  newline='')

        global_file_obj.write('{},{},{}\n'.format(str(sum_of_all_NOK_source_files_per_all_modules), args.build_number, args.data ))
        global_file_obj.close()

        with open(os.path.join(output_path, 'Peer_Review_and_warningReport_analysis_history.csv'), newline='') as f:
            reader = csv.reader(f)
            csv_list = list(reader)

        plot_statistics_per_run(csv_list= csv_list, output_path=output_path )
        plot_statistics_per_module(modules = modules_list_for_plot[0:], list_of_number_of_not_ok_source_files_per_module = list_of_number_of_not_ok_source_files_per_module, output_path=output_path)
    else:
        os.system('echo No Build Number and Date has been found! Plots will not be generated!')

    wb.save(os.path.join(output_path, 'Peer_Review_and_warningReport_analysis.xlsx'))

if __name__ == "__main__":
    main()