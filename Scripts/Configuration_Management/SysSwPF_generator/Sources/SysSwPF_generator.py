"""SysSwPF_generator.py: Used to generate summary report based on excel ( RA, etc. ) as input
Parameters has been given from argparse command line interface.
As input, it is needed an paths to csv files and to template
As output, it is needed a path to generate an xlsx file based upon given template
"""

__author__ = 'Ardeleanu Lucian'
__copyright__ = "Copyright 2021, Autoliv Electronic"
__version__ = "1.7"
__email__ = 'lucian.ardeleanu@autoliv.com'
__status__ = 'Released'

# ============ IMPORTS ====================
import argparse
import os

# ==========================================================================
# FUNCTION USED TO EDIT AN EXCEL FILE
# INPUT ARGUMENT: STRING: PATH TO EXCEL FILE, INTEGERS: FIRST ROW, LAST ROW
#                 STRING: SHEET TO BE MODIFIED
# OUTPUT ARGUMENT: NONE ( IT SAVES MODIFIED EXCEL FILE )
# ==========================================================================
def modify_excel_file(first_row, last_row,  sheet_to_work, file_path ):
    #imports
    import openpyxl

    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_to_work]

    sheet.delete_rows(first_row + 4, last_row - first_row - 4 )

    book.save(file_path)
    book.close()

# ==========================================================================
# FUNCTION USED TO FIND IN AN EXCEL FILE
# INPUT ARGUMENT: STRING: PATH TO EXCEL FILE, STRINGS: 2 STRINGS TO EXTRACT
#                 IN BETWEEN, STRING: SHEET IN WHERE TO SEARCH
# OUTPUT ARGUMENT: INTEGER: ROW NUMBERS WHEHRE FOUNDED DATA IS
# ==========================================================================
def check_in_excel(string_to_check_first, string_to_check_last,sheet_to_work, file_path ):
    #imports
    import openpyxl

    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_to_work]

    first_row_to_find = 0
    last_row_to_find = 0

    for row in sheet.iter_rows(min_row = 1, min_col = 1, max_row = sheet.max_row, max_col = 1):
        for cell in row:
            if cell.value == string_to_check_first:
                first_row_to_find = cell.row

    for row in sheet.iter_rows(min_row = 1, min_col = 1, max_row = sheet.max_row, max_col = 1):
        for cell in row:
            if cell.value == string_to_check_last:
                last_row_to_find = cell.row
    return last_row_to_find - first_row_to_find, first_row_to_find, last_row_to_find

# ==========================================================================
# FUNCTION USED TO INSERT LINE IN AN EXCEL FILE
# INPUT ARGUMENT: STRING: PATH TO EXCEL FILE, INTEGERS: NUMBER OF LINE
#                 INTEGER: A PARAMETER WHERE TO INSERT, STRING: PATH TO GENERATE OUTPUT FILE
# OUTPUT ARGUMENT: NONE ( IT GENERATES AN EXCEL FILE BASED ON GIVEN PATH )
# ==========================================================================
def insert_lines(number_of_lines, input_file,first_row_with_text, output_file):

    import openpyxl
    from copy import copy

    book = openpyxl.load_workbook(input_file)
    sheet = book['Documentation']

    # define where to start insert rows
    start_row_to_insert = first_row_with_text + 4
    # Insert rows
    for count_rows in range(1, number_of_lines ):
        sheet.insert_rows(start_row_to_insert)


    # Copy cell style to another cells
    for row in range(0, number_of_lines - 1):
        for column in range(1,5):
            old_cell = sheet.cell(row = start_row_to_insert - 1, column=column)
            new_cell = sheet.cell(row=row + start_row_to_insert, column=column)

            if old_cell.has_style:
                new_cell.font = copy(old_cell.font)
                new_cell.border = copy(old_cell.border)
                new_cell.fill = copy(old_cell.fill)
                new_cell.number_format = copy(old_cell.number_format)
                new_cell.protection = copy(old_cell.protection)
                new_cell.alignment = copy(old_cell.alignment)

    # Insert an empty row between updated table and next section in Excel
    sheet.insert_rows(start_row_to_insert + number_of_lines - 1 )

    # Debug purposes
    #book.save('test____1.xlsx')
    book.save(output_file)
    book.close()

    # OLD IMPLEMENTATION WITH WIN32COM, NOT WORKING ON JENKINS

    # imports
    # from win32com import client
    #
    # # Code
    # xl = client.Dispatch("Excel.Application")
    # wb = xl.Workbooks.Open(input_file)
    # xl.Visible = 1
    # ws = wb.Sheets("Documentation")  # you can put the sheet number instead of it's name
    # start_row_to_insert = first_row_with_text + 3
    # first_text_range = 'A' + str(start_row_to_insert) + ':' + 'D' + str(start_row_to_insert)
    # second_text_range = 'A' + str(start_row_to_insert - 1) + ':' + 'D' + str(start_row_to_insert - 1)
    # for i in range(1, number_of_lines):
    #     x = ws.Range(first_text_range)
    #     x.EntireRow.Insert()
    # for i in range(1, number_of_lines + 1):
    #     ws.Range(second_text_range).Copy()
    #     insert_to_parameter = 'A' + str(start_row_to_insert)
    #     start_row_to_insert += 1
    #     ws.Paste(ws.Range(insert_to_parameter))
    # xl.ActiveWorkbook.SaveAs(output_file)
    # xl.Quit()

# ===============================================================
# FUNCTION USED TO OPEN AND READ CSV FILE INTO A LIST
# INPUT ARGUMENT: STRING: PATH TO CSV FILE
# OUTPUT ARGUMENT: LIST OF STRINGS WITH DATA
# ===============================================================
def append_csv_to_list(input_csv_path):
    #imports
    import csv

    #Init variables and arrays
    data_from_rows = []

    #Code
    with open(input_csv_path, 'rt')as f:
        data = csv.reader(f)
        next(csv.reader(f))
        for row in data:
            data_from_rows.append(row)

    #Returns
    return data_from_rows

# ===============================================================
# FUNCTION USED TO READ DATA FROM EXCEL FILE
# INPUT ARGUMENT: INTEGERS: DELIMITERS BETWEEN DATA TO BE READ,
#                  STRING: SHEET NAME, STRING: PATH TO EXCEL FILE
# OUTPUT ARGUMENT: LIST OF STRINGS WITH DATA
# ===============================================================
def read_from_excel(first_row, first_column, sheet_to_work, file_path):
    #imports
    import openpyxl

    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_to_work]


    row_array = []
    for row in sheet.iter_rows(min_row = first_row, min_col = first_column, max_row = sheet.max_row, max_col = 6): # condition for every row in sheet
        cell_array = []
        for cell in row:
            if cell.value is None:
                cell.value = ''
            cell_array.append(str(cell.value))
        row_array.append(cell_array)


    #Returns
    book.save(file_path)
    book.close()

    return row_array

# ===============================================================
# FUNCTION USED TO WRITE DATA FROM EXCEL FILE
# INPUT ARGUMENT: DELIMITERS WHERE TO WRITE, LIST OF DATA, AND
#                   FILE PATH AND SHEET
# OUTPUT ARGUMENT: NONE
# ===============================================================
def write_in_excel(first_row, first_column, last_row, last_column, array_list, sheet_to_work, file_path):
    #imports
    import openpyxl

    #Init Variables
    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_to_work]


    for row in sheet.iter_rows(min_row = first_row, min_col = first_column, max_row = last_row, max_col = last_column): # condition for every row in sheet
        for cell in row:  # condition used for writing in every cell
            cell.value = array_list[cell.row - first_row][cell.col_idx - first_column];


    #Returns
    book.save(file_path)
    book.close()

# ===============================================================
# FUNCTION USED TO WRITE DATA FROM EXCEL FILE
# INPUT ARGUMENT: DELIMITERS WHERE TO WRITE, LIST OF DATA, AND
#                   FILE PATH AND SHEET
# OUTPUT ARGUMENT: NONE
# ===============================================================
def write_in_excel_entire_file(first_row, first_column, array_list, sheet_to_work, file_path):
    #imports
    import openpyxl

    #Init Variables
    book = openpyxl.load_workbook(file_path)
    sheet = book[sheet_to_work]


    for row in sheet.iter_rows(min_row = first_row, min_col = first_column, max_row = sheet.max_row, max_col = 6): # condition for every row in sheet
        for cell in row:  # condition used for writing in every cell
            cell.value = array_list[cell.row - first_row][cell.col_idx - first_column];


    #Returns
    book.save(file_path)
    book.close()

# ===============================================================
# MAIN FUNCTION OF SCRIPT
# INPUT ARGUMENT: NONE
# OUTPUT ARGUMENT: NONE
# ===============================================================
def main():

    #Init Variables
    ptc_list_to_be_written_in_excel = []
    number_of_rows_to_identify = 0
    first_row = 0
    last_row = 0


    # Arguments for batch file init
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input-ptc-csv', help="Path to the input PTC .csv file. E.G. ...Items Status PTC.csv", required=True)
    parser.add_argument('-l', '--input-plm-csv', help="Path to the input PLM .csv file. E.G. ...Items Status PLM.csv", required=True)
    parser.add_argument('-t', '--input-template', help="Path to the input template excel file. E.G. Template_file.xlsx OR FILE TO UPDATE",required=True)
    parser.add_argument('-o', '--output-excel', help="Path to the outpt excel to fill. E.G  SW Work-Product Follow-up.xlsx", required=True)
    args = parser.parse_args()


    number_of_rows_to_identify, first_row, last_row = check_in_excel(string_to_check_first='SW Development documents', string_to_check_last='Other documents', sheet_to_work = 'Documentation', file_path = args.input_template)

    print(first_row)
    if ( number_of_rows_to_identify == 4 ):

    # ---------------PTC Data Analisys and Work-out ------------------------------------

        # Append all .csv file in a very big list
        ptc_list = append_csv_to_list(os.path.abspath(args.input_ptc_csv))

        # Create lines in excel file to write it properly
        insert_lines(number_of_lines=len(ptc_list) - 1, input_file=args.input_template, first_row_with_text = first_row ,output_file=args.output_excel)

        #Create list to write in excel from ptc .csv file
        temp_list = ['', '','','']
        for row_ptc in ptc_list:
            temp_list= [str(row_ptc[0]) + ' - ' + str(row_ptc[4]), str(row_ptc[1]), str(row_ptc[2]), str(row_ptc[3]) ]
            ptc_list_to_be_written_in_excel.append(temp_list)


        # Update excel file with fresh data from ptc
        write_in_excel(first_row=first_row + 2, first_column=1, last_row=first_row + 2 + len(ptc_list_to_be_written_in_excel) - 1, last_column=4, array_list= ptc_list_to_be_written_in_excel, sheet_to_work = 'Documentation', file_path = args.output_excel)

    else:
        # ---------------PTC Data Analisys and Work-out ------------------------------------

        # Append all .csv file in a very big list
        ptc_list = append_csv_to_list(os.path.abspath(args.input_ptc_csv))

        # Modify excel file
        modify_excel_file(first_row=first_row, last_row=last_row, sheet_to_work = 'Documentation', file_path = args.input_template)

        # Create lines in excel file to write it properly
        insert_lines(number_of_lines=len(ptc_list) - 1, input_file=args.input_template, first_row_with_text = first_row ,output_file=args.output_excel)

        #Create list to write in excel from ptc .csv file
        temp_list = ['', '','','']
        for row_ptc in ptc_list:
            temp_list= [str(row_ptc[0]) + ' - ' + str(row_ptc[4]), str(row_ptc[1]), str(row_ptc[2]), str(row_ptc[3]) ]
            ptc_list_to_be_written_in_excel.append(temp_list)

        # Update excel file with fresh data from ptc
        write_in_excel(first_row=first_row + 2, first_column=1, last_row=first_row + 2 + len(ptc_list_to_be_written_in_excel) - 1, last_column=4, array_list= ptc_list_to_be_written_in_excel, sheet_to_work = 'Documentation', file_path = args.output_excel)

    # ---------------PLM Data Analisys and Work-out -------------------------------------

    # Append all .csv file in a very big list
    plm_list = append_csv_to_list(os.path.abspath(args.input_plm_csv))

    # Append all .xlsx file in a very big list
    excel_list = read_from_excel(first_row = 1, first_column = 1, sheet_to_work = 'Documentation', file_path = args.output_excel)

    # Update excel list with fresh data from .csv list
    for row_excel in excel_list:
        for row_plm in plm_list:
            if (row_plm[5] in row_excel[1]) or (row_excel[1] == row_plm[5]) or (row_plm[5] in row_excel[2]) or (row_excel[2] == row_plm[5]):
                row_excel[0] = row_plm[4]
                row_excel[1] = row_plm[5]
                row_excel[2] = row_plm[14]
                row_excel[3] = row_plm[8]
                row_excel[4] = row_plm[15]
                row_excel[5] = row_plm[6]

    # Write the modified excel list back again in excel ( overwrite i think existed data? )
    write_in_excel_entire_file(first_row = 1, first_column = 1, array_list= excel_list, sheet_to_work = 'Documentation', file_path = args.output_excel)

if __name__ == '__main__':
    main()
