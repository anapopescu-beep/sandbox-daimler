"""ptc_issue_report.py: Used to launch PTC and check CP contents"""

__author__ = 'Dan Dustinta'
__copyright__ = "Copyright 2018, Autoliv Electronic"
__version__ = "1.0.0"
__email__ = 'dan.dustinta@autoliv.com'
__status__ = 'Released'

import argparse
import sys
import os
import io
import re
import subprocess
from openpyxl import Workbook
from openpyxl import load_workbook
import csv
import openpyxl


currentDir = os.getcwd()
sys.path.append(currentDir)


def matchFile(inputLine):
    matchRet = re.match("([0-9]+)\s*(Working Issue|Change Request)\s*(\([^\(\)]*\))\s*([^\(\)]*\([^\(\)]*\))\s*(.*)\s*(\[.*)", inputLine)
    return matchRet

def matchSoftIngName(inputLine, numberOfName):
   if numberOfName == 2:
       matchRet = re.match("([^\(\)]*\([^\(\)]*\))\s*([^\(\)]*\([^\(\)]*\))", inputLine)
   elif numberOfName == 3:
       matchRet = re.match("([^\(\)]*\([^\(\)]*\))\s*([^\(\)]*\([^\(\)]*\))\s*([^\(\)]*\([^\(\)]*\))", inputLine)
   elif numberOfName == 4:
        matchRet = re.match("([^\(\)]*\([^\(\)]*\))\s*([^\(\)]*\([^\(\)]*\))\s*([^\(\)]*\([^\(\)]*\))\s*([^\(\)]*\([^\(\)]*\))", inputLine)
   return matchRet

def run_command_str(commandStr):
    printStr = []

    p = subprocess.Popen(commandStr, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

    for line in io.TextIOWrapper(p.stdout):
        line = line.strip()
        print_allowed = True
        if line:
            printStr.append(line)

    for line in io.TextIOWrapper(p.stderr):
        line = line.strip()
        print_allowed = True
        if line:
            printStr.append(line)

    return printStr


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-n', '--hostname', help="PTC Host Name", required=True)
    parser.add_argument('-u', '--username', help="PTC User Name", required=True)
    parser.add_argument('-p', '--password', help="PTC User Password", required=True)
    parser.add_argument('-q', '--query', help="PTC Query", required=True)
    parser.add_argument('-o', '--output-path', help="Path to the output directory", required=False)
    parser.add_argument('-t', '--template-path', help="Path to the template directory", required=False)

    args = parser.parse_args()
    commandPrintRet = run_command_str(
        'im issues --fields=ID,"Issue Type",State,"Created By","Assignee Analyze","Assignee SW","Assignee Integrator","Restraints_Assignee_Validation","Issue Title" --query="{PTC_Query}"  --sortAscending '
        '--hostname={hostname} --port=7001 --password={password} --user={username}'.format(
            hostname=args.hostname,
            password=args.password,
            username=args.username,
            PTC_Query=args.query
        ))

    dict = {}
    output_file_str = os.path.normpath(os.path.join(args.output_path, 'IssueReport.xlsx'))

    template_path=args.template_path
    book = openpyxl.load_workbook(template_path)
    sheet = book.active

    # Copy the content of file if exist
    if os.path.exists(output_file_str):
        wbInput = load_workbook(filename=output_file_str, data_only=True)
        for wbSheet in wbInput.sheetnames:
            currSheet = wbInput[wbSheet]
            for col in currSheet.iter_rows(min_row=2):
                id=True
                lastID=" "
                for cell in col:
                    if id==True:
                        id=False
                        lastID = cell.value
                        if cell.value not in dict:
                            dict[str(cell.value)]=[]
                    else:
                        dict[str(lastID)].append(cell.value)

    for line in commandPrintRet:
        testvar = matchFile(line)
        if testvar is not None:
            issueFound = False
            for issue, data in dict.items():
                if issue != None:
                    if issue == testvar.group(1):
                        dict[issue][0] = testvar.group(2)
                        dict[issue][1] = testvar.group(3)
                        dict[issue][2] = testvar.group(4)
                        dict[issue][3] = testvar.group(5)
                        dict[issue][4] = testvar.group(6)
                        issueFound = True
                        break
            noNames = testvar.group(5).count('(')
            if 'Implemented' in testvar.group(3):
                swdNames = matchSoftIngName(testvar.group(5), noNames)
                name = swdNames.group(1)
            elif 'Reviewed' in testvar.group(3):
                swdNames = matchSoftIngName(testvar.group(5), noNames)
                name = swdNames.group(3)
            elif 'Verified' in testvar.group(3):
                swdNames = matchSoftIngName(testvar.group(5), noNames) 
                name = swdNames.group(noNames)
            elif 'Closed' in testvar.group(3):
                swdNames = matchSoftIngName(testvar.group(5), noNames)
                name = swdNames.group(noNames)

            if issueFound == False:
                dict[testvar.group(1)] = [testvar.group(2), testvar.group(3), testvar.group(4), name, testvar.group(6)]
            if len(dict[testvar.group(1)]) > 5:
                    sheet.append([testvar.group(1), testvar.group(2), testvar.group(3), testvar.group(4),
                                 name, testvar.group(6), dict[testvar.group(1)][5], dict[testvar.group(1)][6], dict[testvar.group(1)][7]])
            else:
                sheet.append([testvar.group(1), testvar.group(2), testvar.group(3), testvar.group(4),
                                 name, testvar.group(6)])

        else:
                print(line)


    book.save(output_file_str)
    book.close()

if __name__ == '__main__':
    main()

################ Configuration ####################
# -n="ALVA-MKS03"
# -p="jnk01"
# -u="Jenkins01"
# -q="ToBeVerifed_for_AUDI_TR6"
# -o=s:/Statistics/DeliveryForSWV
# -t=s:/Statistics/DeliveryForSWV/Template/IssueReport_Template.xlsx